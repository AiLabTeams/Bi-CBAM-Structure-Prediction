import xlrd
import openpyxl
import numpy as np
import torch.nn as nn
import torchvision.transforms as transforms
import torch
import xlwt
from torchvision import transforms
import torch.utils.data as Data
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import StandardScaler,MinMaxScaler
from sklearn.metrics import r2_score
from  matplotlib import pyplot as plt
from transformers import ViTModel
import os
from PIL import Image
import matplotlib.image as mpimg
import torch.nn.functional as F
from math import exp
import numpy as np
import timm
import numpy as np
import pandas as pd
import random
from torch.utils.data import DataLoader, TensorDataset
from torchvision import transforms
from sklearn.model_selection import KFold

seed = 42
torch.manual_seed(seed)
np.random.seed(seed)
random.seed(seed)
#check if CUDA is available
use_cuda=torch.cuda.is_available()
print("cuda:",use_cuda)
if torch.cuda.is_available():
    torch.cuda.manual_seed_all(seed)
device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')

path = ("/home/dell/Downloads/attention-cnn/result.xlsx")
df = pd.read_excel(path)
ranges = []
for column in ['Times', 'IA', 'SphereRad', 'OADphi1', 'OADphi2', 'OADphi3']:
    min_val = df[column].min()
    max_val = df[column].max()
    ranges.append([min_val, max_val])
ranges = torch.tensor(ranges, dtype=torch.float32)
print("标签范围 (Ranges):", ranges)

label_dict = {}
for i, row in df.iterrows():
    image_name = str(row['ImageName'])
    label_values = [row['Times'], row['IA'], row['SphereRad'], row['OADphi1'], row['OADphi2'], row['OADphi3']]
    label_dict[image_name] = label_values

pictur_list = []
label_list=[]
# loader=transforms.Compose([transforms.RandomHorizontalFlip(),
#     transforms.RandomRotation(10),
#     transforms.ColorJitter(brightness=0.2, contrast=0.2, saturation=0.2),
#     transforms.ToTensor(),
#     transforms.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5))])
loader=transforms.Compose([
    transforms.ToTensor(),
    transforms.Normalize((0.5, 0.5, 0.5), (0.5, 0.5, 0.5))])
unloader=transforms.ToPILImage()
image_path="/home/dell/Downloads/attention-cnn/123-gray-zip/"


def extract_number(file_name):
    try:
        return int(os.path.splitext(file_name)[0])
    except ValueError:
        return float('inf')  # 如果文件名不含数字，将其放到排序最后

files = sorted(os.listdir(image_path), key=extract_number)
selected_files=files
# selected_files=files[:9896]
# print(selected_files)

for file in selected_files:

    name, ext = os.path.splitext(file)
    if ext.lower() not in ['.jpg', '.png', '.jpeg']:
        continue
    # 根据图片名称获取标签
    if name in label_dict:
        label = label_dict[name]
        label=label[:6]
    else:
        # 若找不到对应标签则跳过或处理异常
        continue

    img = Image.open(os.path.join(image_path, file)).convert('RGB')
    img = loader(img)
    pictur_list.append(img)
    label_list.append(label)

picture_tensor = torch.stack(pictur_list)
label_tensor1 = torch.tensor(label_list, dtype=torch.float32)
def normalize_labels(label_tensor, ranges):
    """
    对标签进行归一化
    :param label_tensor: 标签张量
    :param ranges: 每个标签的取值范围 (max_value - min_value)
    :return: 归一化后的标签张量
    """
    return (label_tensor - ranges[:, 0]) / (ranges[:, 1] - ranges[:, 0])

# 定义标签的取值范围，例如:

# ranges = torch.tensor([
#     [0, 3],    # 标签 1 范围
#     [0, 90],   # 标签 2 范围
#     [0, 225],   # 标签 3 范围
#     [0, 360],     # 标签 4 范围
#     [0, 360],   # 标签 5 范围
# ], dtype=torch.float32)

# 转换范围为 PyTorch 张量
ranges = ranges.to(label_tensor1.device)

# 对标签进行归一化
label_tensor= normalize_labels(label_tensor1, ranges)
print("归一化后的标签形状:", label_tensor.shape)
print("归一化后的标签范围:", label_tensor.min(dim=0).values, label_tensor.max(dim=0).values)
print("归一化后的标签示例:", label_tensor[:6])

# 数据集划分
total_count = len(picture_tensor)
train_ratio = 0.75
val_ratio = 0.15
test_ratio = 0.10

# 计算每个数据集的数量
train_count = int(total_count * train_ratio)
val_count = int(total_count * val_ratio)
test_count = total_count - train_count - val_count

# 数据集划分
train_data1 = picture_tensor[:train_count]
train_label1 = label_tensor[:train_count]

val_data1= picture_tensor[train_count:train_count + val_count]
val_label1 = label_tensor[train_count:train_count + val_count]

test_data = picture_tensor[train_count + val_count:]
test_label = label_tensor[train_count + val_count:]


# 至此完成数据的加载与划分
print("训练集大小:", train_data1.shape, train_label1.shape)
print("验证集大小:", val_data1.shape, val_label1.shape)
print("测试集大小:", test_data.shape, test_label.shape)


# 计算一维的高斯分布向量
def gaussian(window_size, sigma):
    gauss = torch.Tensor([exp(-(x - window_size // 2) ** 2 / float(2 * sigma ** 2)) for x in range(window_size)])
    return gauss / gauss.sum()


# 创建高斯核，通过两个一维高斯分布向量进行矩阵乘法得到
# 可以设定channel参数拓展为3通道
def create_window(window_size, channel=1):
    _1D_window = gaussian(window_size, 1.5).unsqueeze(1)
    _2D_window = _1D_window.mm(_1D_window.t()).float().unsqueeze(0).unsqueeze(0)
    window = _2D_window.expand(channel, 1, window_size, window_size).contiguous()
    return window


# 计算SSIM
# 直接使用SSIM的公式，但是在计算均值时，不是直接求像素平均值，而是采用归一化的高斯核卷积来代替。
# 在计算方差和协方差时用到了公式Var(X)=E[X^2]-E[X]^2, cov(X,Y)=E[XY]-E[X]E[Y].
# 正如前面提到的，上面求期望的操作采用高斯核卷积代替。
def ssim(img1, img2, window_size=11, window=None, size_average=True, full=False, val_range=None):
    # Value range can be different from 255. Other common ranges are 1 (sigmoid) and 2 (tanh).
    if val_range is None:
        if torch.max(img1) > 128:
            max_val = 255
        else:
            max_val = 1

        if torch.min(img1) < -0.5:
            min_val = -1
        else:
            min_val = 0
        L = max_val - min_val
    else:
        L = val_range

    padd = 0
    (_, channel, height, width) = img1.size()
    if window is None:
        real_size = min(window_size, height, width)
        window = create_window(real_size, channel=channel).to(img1.device)

    mu1 = F.conv2d(img1, window, padding=padd, groups=channel)
    mu2 = F.conv2d(img2, window, padding=padd, groups=channel)

    mu1_sq = mu1.pow(2)
    mu2_sq = mu2.pow(2)
    mu1_mu2 = mu1 * mu2

    sigma1_sq = F.conv2d(img1 * img1, window, padding=padd, groups=channel) - mu1_sq
    sigma2_sq = F.conv2d(img2 * img2, window, padding=padd, groups=channel) - mu2_sq
    sigma12 = F.conv2d(img1 * img2, window, padding=padd, groups=channel) - mu1_mu2

    C1 = (0.01 * L) ** 2
    C2 = (0.03 * L) ** 2

    v1 = 2.0 * sigma12 + C2
    v2 = sigma1_sq + sigma2_sq + C2
    cs = torch.mean(v1 / v2)  # contrast sensitivity

    ssim_map = ((2 * mu1_mu2 + C1) * v1) / ((mu1_sq + mu2_sq + C1) * v2)

    if size_average:
        ret = ssim_map.mean()
    else:
        ret = ssim_map.mean(1).mean(1).mean(1)

    if full:
        return ret, cs
    return ret


# Classes to re-use window
class SSIM(torch.nn.Module):
    def __init__(self, window_size=11, size_average=True, val_range=None):
        super(SSIM, self).__init__()
        self.window_size = window_size
        self.size_average = size_average
        self.val_range = val_range

        # Assume 1 channel for SSIM
        self.channel = 1
        self.window = create_window(window_size)

    def forward(self, img1, img2):
        (_, channel, _, _) = img1.size()

        if channel == self.channel and self.window.dtype == img1.dtype:
            window = self.window
        else:
            window = create_window(self.window_size, channel).to(img1.device).type(img1.dtype)
            self.window = window
            self.channel = channel

        return ssim(img1, img2, window=window, window_size=self.window_size, size_average=self.size_average)
class SEBlock(nn.Module):
    """
    通道注意力机制模块 (Squeeze-and-Excitation Block)
    """
    def __init__(self, kernel_size=7):
        super(SEBlock, self).__init__()
        assert kernel_size % 2 == 1, "Kernel size must be odd for padding calculation"
        padding = kernel_size // 2
        self.conv1 = nn.Conv2d(2, 1, kernel_size=kernel_size, padding=padding, bias=False)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        avg_out = torch.mean(x, dim=1, keepdim=True)
        max_out, _ = torch.max(x, dim=1, keepdim=True)
        cat = torch.cat([avg_out, max_out], dim=1)
        out = self.conv1(cat)
        return x * self.sigmoid(out)

class CAB(nn.Module):
    def __init__(self, in_channel, ratio=16):
        super(CAB, self).__init__()
        reduced_channels = max(1, in_channel // ratio)  # 设置最小通道数为 1
        self.avg_pool = nn.AdaptiveAvgPool2d(1)
        self.max_pool = nn.AdaptiveMaxPool2d(1)
        self.fc1 = nn.Conv2d(in_channel, reduced_channels, kernel_size=1, bias=False)
        self.relu1 = nn.ReLU()
        self.fc2 = nn.Conv2d(reduced_channels, in_channel, kernel_size=1, bias=False)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        avg_out = self.fc2(self.relu1(self.fc1(self.avg_pool(x))))
        max_out = self.fc2(self.relu1(self.fc1(self.max_pool(x))))
        out = avg_out + max_out
        return x * self.sigmoid(out)

class CBAM(nn.Module):
    def __init__(self, in_channels, reduction=16, kernel_size=7):
        super(CBAM, self).__init__()
        self.channel_attention = CAB(in_channels, reduction)
        self.spatial_attention = SEBlock(kernel_size)

    def forward(self, x):
        channel_weight = self.channel_attention(x)
        spatial_weight = self.spatial_attention(x)
        # print("Channel weight range:", channel_weight.min().item(), channel_weight.max().item())
        # print("Spatial weight range:", spatial_weight.min().item(), spatial_weight.max().item())
        # x = x * channel_weight
        # x = x * spatial_weight
        x = x + 0.1 * self.channel_attention(x)
        x = x + 0.2 * self.spatial_attention(x)
        return x

class CNN(nn.Module):
    def __init__(self, kernel_size=3):
        super(CNN, self).__init__()
        self.conv = nn.Sequential(
            nn.Conv2d(3, 32, kernel_size, padding=1),
            nn.BatchNorm2d(32),
            nn.ReLU(),
            CBAM(32),
            # SEBlock(32),  # 通道注意力机制

            nn.Conv2d(32, 64, kernel_size, padding=1),
            nn.BatchNorm2d(64),
            nn.ReLU(),
            CBAM(64),
            # SEBlock(64),  # 通道注意力机制

            nn.Conv2d(64, 128, kernel_size, padding=1),
            nn.BatchNorm2d(128),
            nn.ReLU(),
            CBAM(128),
            # SEBlock(128),  # 通道注意力机制

            nn.AdaptiveAvgPool2d((8, 8))  # 动态池化到固定尺寸
        )
        self.fc = nn.Sequential(
            nn.Flatten(),
            nn.Linear(128 * 8 * 8, 256),
            nn.BatchNorm1d(256),
            nn.ReLU(),
            nn.Dropout(0.5),
            nn.Linear(256, 128),
            nn.BatchNorm1d(128),
            nn.ReLU(),
            nn.Dropout(0.5),
            nn.Linear(128, 6)  # 输出 6 个制备参数
        )

    def forward(self, x):
        x = self.conv(x)
        x = self.fc(x)
        return x
class ForwardCNN(nn.Module):
    def __init__(self, kernel_size=3):
        super(ForwardCNN, self).__init__()
        self.fc = nn.Sequential(
            nn.Linear(6, 128),
            nn.BatchNorm1d(128),
            nn.ReLU(),
            nn.Dropout(0.5),
            nn.Linear(128, 256),
            nn.BatchNorm1d(256),
            nn.ReLU(),
            nn.Dropout(0.5),
            nn.Linear(256, 36 * 18 * 10),
            nn.BatchNorm1d(36 * 18 * 10),
            nn.ReLU()
        )
        self.deconv = nn.Sequential(
            nn.ConvTranspose2d(36, 24, kernel_size=3, stride=2, padding=1, output_padding=(1,0)),
            nn.BatchNorm2d(24),
            nn.ReLU(),
            CBAM(24),
            # SEBlock(24),  # 通道注意力机制

            nn.ConvTranspose2d(24, 16, kernel_size=3, stride=2, padding=1, output_padding=(1,1)),
            nn.BatchNorm2d(16),
            nn.ReLU(),
            CBAM(16),
            # SEBlock(16),  # 通道注意力机制

            nn.ConvTranspose2d(16, 3, kernel_size=3, stride=2,padding=1, output_padding=(1,0)),
            nn.ReLU()  # 输出范围映射到 [-1, 1]
        )

    def forward(self, x):
        x = self.fc(x)
        x = x.view(-1, 36, 18, 10)  # 恢复为卷积张量形状
        x = self.deconv(x)
        # print(x.shape)
        x = F.interpolate(x, size=(138, 80), mode='bilinear', align_corners=False)
        # print(x.shape)
        return x


model_inverse = CNN()
model_forward=ForwardCNN()
# model_inverse=torch.load("model/model_inverse_l1.pkl")
# model_forward=torch.load("model/model_forward_l1.pkl")
model_inverse=model_inverse.to(device)
model_forward=model_forward.to(device)
# loss_fun = torch.nn.MSELoss()
# optimizer = torch.optim.Adam(model.parameters(), lr=0.0005)
train_set = Data.TensorDataset(train_data1,train_label1)
val_set=Data.TensorDataset(val_data1,val_label1)
test_set=Data.TensorDataset(test_data,test_label)
#
# numbers="Bi-CBAM-"+str(1)
# epochs = 1000
# k_folds = 10
# kf = KFold(n_splits=k_folds, shuffle=True, random_state=42)
#
# # 存储每折的结果
# fold_results = []
# training_results = []
# # 开始交叉验证
# for fold, (train_idx, val_idx) in enumerate(kf.split(picture_tensor)):
#     print(f"Fold {fold + 1}/{k_folds}")
#
#     # 获取训练集和验证集
#     train_data, train_label = picture_tensor[train_idx], label_tensor[train_idx]
#     val_data, val_label = picture_tensor[val_idx], label_tensor[val_idx]
#
#     # 数据加载器
#     dataiter_train = torch.utils.data.DataLoader(
#         dataset=torch.utils.data.TensorDataset(train_data, train_label),
#         batch_size=800,  # 调整为合适的批量大小
#         shuffle=True
#     )
#     dataiter_val = torch.utils.data.DataLoader(
#         dataset=torch.utils.data.TensorDataset(val_data, val_label),
#         batch_size=800,
#         shuffle=False
#     )
#
#     # 定义损失函数和优化器
#     loss_fun1 = torch.nn.MSELoss()
#     loss_fun2 = torch.nn.MSELoss()
#     opt1 = torch.optim.Adam(model_inverse.parameters(), lr=0.01, weight_decay=1E-5)
#     opt2 = torch.optim.Adam(model_forward.parameters(), lr=0.01, weight_decay=1E-5)
#     scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(opt1, mode='min', factor=0.5, patience=10)
#
#     # 存储训练结果
#     best_r2 = -float('inf')
#
#     for epoch in range(epochs):
#         # 模型训练阶段
#         model_inverse.train()
#         model_forward.train()
#         train_losses = []
#         train_r2_scores = []
#
#         for x, y in dataiter_train:
#             x, y = x.to(device), y.to(device)
#
#             # 模型 1：从图片到标签
#             opt1.zero_grad()
#             preds = model_inverse(x)
#             loss_inverse = loss_fun1(preds, y)
#
#             # 模型 2：从标签到图片
#             opt2.zero_grad()
#             recon_images = model_forward(preds)
#             loss_forward = loss_fun2(recon_images, x)
#
#             # 总损失
#             total_loss = loss_inverse * 0.5 + loss_forward * 0.5
#             total_loss.backward()
#
#             opt1.step()
#             opt2.step()
#
#             train_losses.append(total_loss.item())
#             train_r2_scores.append(r2_score(y.cpu().detach().numpy(), preds.cpu().detach().numpy()))
#
#         scheduler.step(np.mean(train_losses))
#
#         # 验证阶段
#         model_inverse.eval()
#         model_forward.eval()
#         val_losses = []
#         val_r2_scores = []
#         ssim_val1 = []
#
#         with torch.no_grad():
#             for x, y in dataiter_val:
#                 x, y = x.to(device), y.to(device)
#
#                 preds = model_inverse(x)
#                 loss_inverse = loss_fun1(preds, y)
#
#                 val_stru = model_forward(preds)
#                 loss_forward = loss_fun2(val_stru, x)
#
#                 total_loss = loss_inverse*0.5 + loss_forward*0.5
#                 val_losses.append(total_loss.item())
#                 val_r2_scores.append(r2_score(y.cpu().detach().numpy(), preds.cpu().detach().numpy()))
#                 ssim_val = ssim(x.cpu().detach(), val_stru.cpu().detach(), size_average=True)
#                 ssim_val1.append(ssim_val)
#
#         avg_train_loss = np.mean(train_losses)
#         avg_train_r2 = np.mean(train_r2_scores)
#         avg_val_loss = np.mean(val_losses)
#         avg_val_r2 = np.mean(val_r2_scores)
#         avg_ssim = np.mean(ssim_val1)
#
#         print(f"Epoch {epoch + 1}/{epochs}, Fold {fold + 1}: Train Loss: {avg_train_loss:.4f}, Train R2: {avg_train_r2:.4f}, "
#               f"Val Loss: {avg_val_loss:.4f}, Val R2: {avg_val_r2:.4f}, Val SSIM: {avg_ssim:.4f}")
#
#         # 保存最优模型
#         if avg_val_r2 > best_r2:
#             best_r2 = avg_val_r2
#             torch.save(model_inverse.state_dict(), f"model/best_model_inverse_fold{fold + 1}.pth")
#             torch.save(model_forward.state_dict(), f"model/best_model_forward_fold{fold + 1}.pth")
#             print(f"New best model saved for Fold {fold + 1} with R2: {best_r2:.4f}")
#
#     # 保存每折验证结果
#     fold_results.append({
#         "fold": fold + 1,
#         "best_val_r2": best_r2
#     })
#
# # 打印交叉验证结果
# print("Cross-Validation Results:")
# for result in fold_results:
#     print(f"Fold {result['fold']} - Best Val R2: {result['best_val_r2']:.4f}")
# avg_r2 = np.mean([result["best_val_r2"] for result in fold_results])
# print(f"Average Best Val R2: {avg_r2:.4f}")

dataiter_train = Data.DataLoader(dataset = train_set,
                           batch_size =1000,
                           num_workers=0,
                           shuffle = True)
dataiter_val = Data.DataLoader(dataset = val_set,
                           batch_size =1000,
                           num_workers=0,
                           shuffle = True)
dataiter_test = Data.DataLoader(dataset = test_set,
                           batch_size =1000,
                           num_workers=0,
                           shuffle = True)
# loss_fun1 = torch.nn.L1Loss()
# loss_fun2 = torch.nn.L1Loss()

loss_fun1 = torch.nn.MSELoss()
loss_fun2 = torch.nn.MSELoss()
opt1= torch.optim.Adam(model_inverse.parameters(),lr=0.01,weight_decay=1E-5)
opt2= torch.optim.Adam(model_forward.parameters(), lr=0.01,weight_decay=1E-5)
scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(opt1, mode='min', factor=0.5, patience=10)
val_loss=[]
val_loss_total=[]
tra_loss=[]
tra_forward_loss=[]
r_2_total=[]
r_2_1=[]
train_r2_scores = []
ssim_val1=[]
ssim_val_total=[]
epochs = 20
best_r2 = -float('inf')
best_model_path = "best_model.pth"
numbers="Bi-CBM-"+str(1)
training_results = []
for epoch in range(epochs):
    model_inverse.train()
    model_forward.train()
    train_losses = []
    train_loss=0

    for x, y in dataiter_train:
        x, y = x.to(device), y.to(device)

        # 模型 1：从图片到标签
        opt1.zero_grad()
        preds = model_inverse(x)
        loss_inverse = loss_fun1(preds, y)

        # 模型 2：从标签到图片
        opt2.zero_grad()
        recon_images = model_forward(preds)
        loss_forward = loss_fun2(recon_images, x)

        # 总损失
        total_loss = loss_inverse*0.5 + loss_forward*0.5
        total_loss.backward()

        opt1.step()
        opt2.step()
        train_loss+=total_loss.item()
        train_losses.append(total_loss.item())
        train_r2_scores.append(r2_score(y.cpu().detach().numpy(), preds.cpu().detach().numpy())
)
    scheduler.step(train_loss)
    # 验证
    model_inverse.eval()
    model_forward.eval()
    val_losses = []
    val_r2_scores = []

    with torch.no_grad():
        for x, y in dataiter_val:
            x, y = x.to(device), y.to(device)

            preds = model_inverse(x)
            loss_inverse = loss_fun1(preds, y)

            val_stru = model_forward(preds)
            loss_forward = loss_fun2(val_stru, x)

            total_loss = loss_inverse + loss_forward
            val_losses.append(total_loss.item())
            val_r2_scores.append(r2_score(y.cpu().detach().numpy(), preds.cpu().detach().numpy()))
            val_stru_cpu = val_stru.cpu()
            ssim_val = ssim(x.cpu().detach(), val_stru_cpu.detach(), size_average=True)
            ssim_val1.append(ssim_val)

    avg_train_loss = np.mean(train_losses)
    avg_train_r2= np.mean(train_r2_scores)
    avg_val_loss = np.mean(val_losses)
    avg_val_r2 = np.mean(val_r2_scores)
    avg_ssim=np.mean(ssim_val1)

    print(f"Epoch {epoch+1}/{epochs}, Train Loss: {avg_train_loss:.4f}, Train r2: {avg_train_r2:.4f},Val Loss: {avg_val_loss:.4f}, Val R2: {avg_val_r2:.4f}, Val SSIM: {avg_ssim:.4f}")

    # 记录结果用于保存
    training_results.append({
        "epoch": epoch + 1,
        "train_loss": avg_train_loss,
        "train_r2": avg_train_r2,
        "val_loss": avg_val_loss,
        "val_r2": avg_val_r2,
        "val_ssim": avg_ssim,
    })
    # 保存最优模型
    if avg_val_r2 > best_r2:
        best_r2 = avg_val_r2
        torch.save(model_inverse, "model/best_model_inverse" + numbers + ".pkl")
        torch.save(model_forward, "model/best_model_forward" + numbers + ".pkl")
        print(f"New best model saved with R2: {best_r2:.4f}")

# 保存最终模型
torch.save(model_inverse,"model/final_model_inverse"+numbers+".pkl")
torch.save(model_forward,"model/final_model_forward"+numbers+".pkl")
print(f"Final model saved")
def save_training_results_to_excel(file_path, results):
    """
    保存训练和验证结果到 Excel 文件
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Training Results'

    # 写入表头
    headers = ['Epoch', 'Train Loss', 'Val Loss', 'R2', 'SSIM']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # 写入数据
    for row, result in enumerate(results, start=2):
        ws.cell(row=row, column=1, value=result['epoch'])
        ws.cell(row=row, column=2, value=result['train_loss'])
        ws.cell(row=row, column=3, value=result['train_r2'])
        ws.cell(row=row, column=4, value=result['val_loss'])
        ws.cell(row=row, column=5, value=result['val_r2'])
        ws.cell(row=row, column=6, value=result['val_ssim'])

    wb.save(file_path)
    print(f"训练结果已保存到: {file_path}")
save_training_results_to_excel("training_results"+numbers+".xlsx", training_results)

test_lable=test_label.to(device)
test_data=test_data.to(device)
test_out = model_inverse(test_data)
test_stru = model_forward(test_out)
test_out=test_out.cpu()
test_lable=test_lable.cpu()
test_stru=test_stru.cpu()
test_data=test_data.cpu()
ssim_test = ssim(test_stru.detach(), test_data.detach(),size_average=True)
print("测试的SSIM距离:{:.4f}".format(ssim_test))

def denormalize_image(tensor, mean, std):
    mean = torch.tensor(mean).view(-1, 1, 1).to(tensor.device)
    std = torch.tensor(std).view(-1, 1, 1).to(tensor.device)
    return tensor * std + mean

# 示例数据
mean = [0.5, 0.5, 0.5]
std = [0.5, 0.5, 0.5]

# 反归一化
test_data_ori = denormalize_image(test_data, mean, std)
test_stru_ori = denormalize_image(test_stru, mean, std)
# 转换为像素范围 [0, 255]
test_data_image = (test_data_ori * 255).clamp(0, 255).byte()
test_stru_image = (test_stru_ori * 255).clamp(0, 255).byte()
true_pic="result/true"+numbers+"/"
pre_pic="result/pre1"+numbers+"/"
toPIL=transforms.ToPILImage()
test_size=len(test_label)
for i in range (test_size):
    pic=toPIL(test_data_image[i])
    pic.save(true_pic+str(i) +'.jpg')
    pic1 = toPIL(test_stru_image[i])
    pic1.save(pre_pic + str(i) + '.jpg')

def denormalize_labels(normalized_tensor, ranges):
    """
    将归一化后的张量反归一化
    :param normalized_tensor: 归一化后的张量
    :param ranges: 每个标签的取值范围 (min, max)
    :return: 反归一化后的张量
    """
    return normalized_tensor * (ranges[:, 1] - ranges[:, 0]) + ranges[:, 0]

# 示例：反归一化
test_label = denormalize_labels(test_label, ranges)
print("反归一化后的标签形状:", test_label.shape)
test_out=denormalize_labels(test_out,ranges)

S=[]
pred_times=test_out[:,0].detach()
pred_azi_angle1=test_out[:,3].detach()
pred_azi_angle2=test_out[:,4].detach()
# pred_azi_angle3=test_out[:,5].detach()
pred_rad=test_out[:,2].detach()
pred_inc_angle=test_out[:,1].detach()
True_inc_angle=test_label[:,1].detach()
True_azi_angle1=test_label[:,3].detach()
True_azi_angle2=test_label[:,4].detach()
# True_azi_angle3=test_lable[:,5].detach()
True_rad=test_label[:,2].detach()
True_times=test_label[:,0].detach()
# val_loss1=np.ravel(val_loss)
# r_2_total1=np.ravel(r_2_total)

#画图
m=100
plt.title("测试")
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.xlabel("样本数")
plt.ylabel("结构参数")
plt.subplot(3, 2, 1)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1),pred_inc_angle[0:m], label='入射角预测')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_inc_angle[0:m], label='入射角真实')
plt.legend()
plt.subplot(3, 2, 2)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_azi_angle1[0:m], label='方位角1预测')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_azi_angle1[0:m], label='方位角1真实')
plt.legend()
plt.subplot(3, 2, 3)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_rad[0:m], label='半径预测')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_rad[0:m], label='半径真实')
plt.legend()
plt.subplot(3, 2, 4)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_azi_angle2[0:m], label='方位角2预测')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_azi_angle2[0:m], label='方位角2真实')
plt.legend()

plt.subplot(3, 2, 5)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_times[0:m], label='Times 预测')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_times[0:m], label='Times 真实')
plt.legend()
# plt.subplot(3, 2, 6)
# plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_azi_angle3[0:m], label='方位角3预测')
# plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_azi_angle3[0:m], label='方位角3真实')
# plt.legend()

plt.show()

def save_results_to_excel(file_path, true_values, pred_values, col_names):
    """
    保存测试结果到 Excel
    :param file_path: 文件路径
    :param true_values: 真实值
    :param pred_values: 预测值
    :param col_names: 列名
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Results'

    # 写入列名
    for index, col in enumerate(col_names):
        ws.cell(row=1, column=index + 1, value=col)

    # 写入数据
    for i in range(len(true_values)):
        for j, value in enumerate(true_values[i] + pred_values[i]):
            ws.cell(row=i + 2, column=j + 1, value=value)

    wb.save(file_path)
    print(f"结果保存到: {file_path}")

# 示例调用
save_results_to_excel(
    file_path="result/test-pred-paramas-"+numbers+".xlsx",
    true_values=[[True_inc_angle, True_azi_angle1,True_azi_angle2, True_rad, True_times]],
    pred_values=[[pred_inc_angle, pred_azi_angle1, pred_azi_angle2,pred_rad, pred_times]],
    col_names=['True_inc', 'True_azi1', 'True_azi2','True_rad', 'True_times', 'Pred_inc', 'Pred_azi1', 'Pred_azi2','Pred_rad', 'Pred_times']
)












