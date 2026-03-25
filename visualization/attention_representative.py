import numpy as np
import torch.nn as nn
import torchvision.transforms as transforms
import torch
import random
try:
    import openpyxl
except ImportError:
    openpyxl = None
import torch.utils.data as Data
try:
    from sklearn.metrics import r2_score
except ImportError:
    def r2_score(y_true, y_pred):
        y_true = np.asarray(y_true, dtype=np.float64)
        y_pred = np.asarray(y_pred, dtype=np.float64)
        ss_res = np.sum((y_true - y_pred) ** 2)
        ss_tot = np.sum((y_true - np.mean(y_true, axis=0, keepdims=True)) ** 2)
        if ss_tot == 0:
            return 0.0
        return float(1.0 - ss_res / ss_tot)
from  matplotlib import pyplot as plt
import os
from PIL import Image
import matplotlib.image as mpimg
import torch.nn.functional as F
from math import exp
import numpy as np
import math
from torch.optim import lr_scheduler
import pandas as pd
from pathlib import Path

#check if CUDA is available
use_cuda=torch.cuda.is_available()
print("cuda:",use_cuda)
device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')


DATA_ROOT = str((PROJECT_ROOT.parent / "data" / "images").resolve())
EXCEL_PATH = str((PROJECT_ROOT.parent / "data" / "labels.xlsx").resolve())
SEED = 42
SOURCE_IMAGE_SIZE = (342, 197)
MODEL_INPUT_SIZE = SOURCE_IMAGE_SIZE
IMAGE_CHANNELS = 3
ENCODER_CHANNELS = 36
ENCODER_H = MODEL_INPUT_SIZE[0] // 2  # one MaxPool2d(2,2)
ENCODER_W = MODEL_INPUT_SIZE[1] // 2
ENCODER_FLAT_DIM = ENCODER_CHANNELS * ENCODER_H * ENCODER_W
STRICT_DETERMINISM = os.getenv("STRICT_DETERMINISM", "0") == "1"
DETERMINISM_WARN_ONLY = not STRICT_DETERMINISM

def set_global_determinism(seed: int = 42, warn_only: bool = True):
    # Python / NumPy / Torch RNG
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)

    # CuDNN deterministic behavior
    if hasattr(torch.backends, "cudnn"):
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False

    # Deterministic algorithms.
    # Some CUDA ops (e.g. adaptive max pool backward) have no deterministic implementation.
    # In warn_only mode, training continues and PyTorch will emit warnings for such ops.
    try:
        torch.use_deterministic_algorithms(True, warn_only=warn_only)
    except TypeError:
        # Compatibility for old PyTorch versions without warn_only argument.
        torch.use_deterministic_algorithms(True)
    except Exception:
        pass

    # Deterministic cublas workspace for CUDA matmul (if CUDA is used)
    os.environ.setdefault("CUBLAS_WORKSPACE_CONFIG", ":4096:8")

set_global_determinism(SEED, warn_only=DETERMINISM_WARN_ONLY)

PROJECT_ROOT = Path(__file__).resolve().parent
OUTPUT_ROOT = PROJECT_ROOT / "visualization_outputs"
MODEL_DIR = PROJECT_ROOT.parent / "models"
TRUE_IMG_DIR = OUTPUT_ROOT / "true"
PRED_IMG_DIR = OUTPUT_ROOT / "pre"
RESULT_DIR = OUTPUT_ROOT / "result"
for _d in [TRUE_IMG_DIR, PRED_IMG_DIR, RESULT_DIR]:
    _d.mkdir(parents=True, exist_ok=True)

DATA_TRANSFORM = transforms.Compose([
    transforms.Resize(SOURCE_IMAGE_SIZE),
    transforms.Resize(MODEL_INPUT_SIZE),
    transforms.Grayscale(num_output_channels=IMAGE_CHANNELS),
    transforms.ToTensor()
])

# Interpretability analysis switches (review-response module, append-only).
# These options do not change model architecture, loss, or prediction outputs.
ENABLE_ATTN_ANALYSIS = True
ATTN_ANALYSIS_ONLY_IN_TEST_MODE = True
# Aggregation over multiple spatial-attention layers.
# "max" is recommended here because it preserves the most selective layer response
# and avoids diluting structure-focused activations by near-uniform shallow layers.
ATTN_AGG_METHOD = "max"  # "mean" or "weighted" or "max"
ATTN_VIS_SAMPLE_IDS = []  # e.g. [0, 5, 20, 33]
ATTN_AUTO_SAMPLES_PER_GROUP = 2
ATTN_EDGE_TOPK = 0.15
ATTN_EDGE_BAND_WIDTH = 3
ATTN_BORDER_IGNORE = 4
ATTN_FOREGROUND_THRESH = 0.08
# Representative-sample selection settings. These are only used for choosing
# qualitative figures; they do not affect the model or the raw attention maps.
ATTN_REP_MIN_FOREGROUND = 0.05
ATTN_REP_MAX_FOREGROUND = 0.35
ATTN_REP_MIN_BG_ATTENTION = 1e-4
ATTN_REP_TOPN_EXPORT = 30
ATTN_SAVE_PER_LAYER = True
ATTN_SAVE_RAW_MAPS = True

ATTN_ROOT = OUTPUT_ROOT / "attention-1"
ATTN_OVERLAY_DIR = ATTN_ROOT / "overlay"
ATTN_PER_LAYER_DIR = ATTN_ROOT / "per_layer"
ATTN_RAW_DIR = ATTN_ROOT / "raw_maps"
for _d in [ATTN_ROOT, ATTN_OVERLAY_DIR, ATTN_PER_LAYER_DIR, ATTN_RAW_DIR]:
    _d.mkdir(parents=True, exist_ok=True)


def canonical_imagename(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
        return "{:g}".format(f)
    except ValueError:
        if s.endswith(".0"):
            return s[:-2]
        return s


def read_labels_from_excel(excel_path):
    try:
        if str(excel_path).lower().endswith(("xlsx", "xlsm", "xls")):
            if openpyxl is not None:
                df = pd.read_excel(excel_path, engine="openpyxl")
            else:
                df = pd.read_excel(excel_path)
        else:
            df = pd.read_excel(excel_path)
    except ImportError as e:
        raise ImportError(
            "Reading Excel requires openpyxl. Please install: pip install openpyxl"
        ) from e
    df.columns = [str(c).strip() for c in df.columns]
    required = ["ImageName", "Times", "IA", "SphereRad", "OADphi1", "OADphi2", "OADphi3"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Excel missing required columns: {missing}")
    df = df.dropna(how="all")
    df["ImageStem"] = df["ImageName"].apply(canonical_imagename)
    return df



def build_samples(data_root, excel_path):
    df = read_labels_from_excel(excel_path)
    img_root = Path(data_root)

    valid_exts = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}
    image_map = {}
    for p in img_root.iterdir():
        if p.is_file() and p.suffix.lower() in valid_exts:
            image_map[p.stem.lower()] = p

    samples = []
    missing_count = 0

    def get_float(v, default):
        try:
            return float(v)
        except Exception:
            return default

    for _, row in df.iterrows():
        stem = row["ImageStem"]
        if not stem:
            missing_count += 1
            continue

        stem_s = str(stem).lower()
        img_path = image_map.get(stem_s)
        if img_path is None:
            for width in range(2, 7):
                padded = stem_s.zfill(width)
                if padded in image_map:
                    img_path = image_map[padded]
                    break

        if img_path is None:
            missing_count += 1
            continue

        times = float(np.clip(get_float(row["Times"], 1.0), 1.0, 3.0))
        ia = float(np.clip(get_float(row["IA"], 0.0), 0.0, 90.0))
        rad = float(np.clip(get_float(row["SphereRad"], 50.0), 50.0, 250.0))

        phi1 = get_float(row["OADphi1"], 0.0) % 360.0
        phi2 = get_float(row["OADphi2"], 0.0) % 360.0
        phi3 = get_float(row["OADphi3"], 0.0) % 360.0
        if times < 2.0:
            phi2 = 0.0
        if times < 3.0:
            phi3 = 0.0

        label = [
            ia / 90.0,
            phi1 / 360.0,
            (rad - 50.0) / 200.0,
            times / 3.0,
            phi2 / 360.0,
            phi3 / 360.0,
        ]

        samples.append((str(img_path), np.asarray(label, dtype=np.float32)))

    if len(samples) == 0:
        raise RuntimeError("No matched images found. Please check DATA_ROOT and EXCEL_PATH.")

    print(f"Dataset matched: {len(samples)}, missing: {missing_count}")
    return samples


class LazyImageLabelDataset(Data.Dataset):
    def __init__(self, samples, transform):
        self.samples = samples
        self.transform = transform

    def __len__(self):
        return len(self.samples)

    def __getitem__(self, idx):
        img_path, label = self.samples[idx]
        with Image.open(img_path) as img:
            img = img.convert("RGB")
            x = self.transform(img)
        y = torch.tensor(label, dtype=torch.float32)
        return x, y


samples = build_samples(DATA_ROOT, EXCEL_PATH)

rng = np.random.RandomState(SEED)
index = np.arange(len(samples))
rng.shuffle(index)

n = len(index)
n_train = int(n * 0.7)
n_val = int(n * 0.2)

train_idx = index[0:n_train].tolist()
val_idx = index[n_train:n_train + n_val].tolist()
test_idx = index[n_train + n_val:].tolist()

full_dataset = LazyImageLabelDataset(samples, transform=DATA_TRANSFORM)
train_set = Data.Subset(full_dataset, train_idx)
val_set = Data.Subset(full_dataset, val_idx)
test_set = Data.Subset(full_dataset, test_idx)


class EarlyStopping(object):
    def __init__(self, monitor: str = 'val_loss', mode: str = 'min', patience: int = 1):
        
        self.monitor = monitor
        self.mode = mode
        self.patience = patience
        self.__value = -math.inf if mode == 'max' else math.inf
        self.__times = 0

    def state_dict(self) -> dict:
        
        return {
            'monitor': self.monitor,
            'mode': self.mode,
            'patience': self.patience,
            'value': self.__value,
            'times': self.__times
        }

    def load_state_dict(self, state_dict: dict):
        
        self.monitor = state_dict['monitor']
        self.mode = state_dict['mode']
        self.patience = state_dict['patience']
        self.__value = state_dict['value']
        self.__times = state_dict['times']

    def __call__(self, metrics) -> bool:
        
        if isinstance(metrics, dict):
            metrics = metrics[self.monitor]

        if (self.mode == 'min' and metrics <= self.__value) or (
                self.mode == 'max' and metrics >= self.__value):
            self.__value = metrics
            self.__times = 0
        else:
            self.__times += 1
        if self.__times >= self.patience:
            return True
        return False


def gaussian(window_size, sigma):
    gauss = torch.Tensor([exp(-(x - window_size // 2) ** 2 / float(2 * sigma ** 2)) for x in range(window_size)])
    return gauss / gauss.sum()


def create_window(window_size, channel=1):
    _1D_window = gaussian(window_size, 1.5).unsqueeze(1)
    _2D_window = _1D_window.mm(_1D_window.t()).float().unsqueeze(0).unsqueeze(0)
    window = _2D_window.expand(channel, 1, window_size, window_size).contiguous()
    return window


def ssim(img1, img2, window_size=11, window=None, size_average=True, full=False, val_range=None):
    # Value range can be different from 255. Other common ranges are 1 (sigmoid) and 2 (tanh).
    if val_range is None:
        if torch.max(img1) > 128:
            max_val = 255
        else:
            max_val = 1

        if torch.min(img1) < -0.5:
            min_val = -1
        else:
            min_val = 0
        L = max_val - min_val
    else:
        L = val_range

    padd = 0
    (_, channel, height, width) = img1.size()
    if window is None:
        real_size = min(window_size, height, width)
        window = create_window(real_size, channel=channel).to(img1.device)

    mu1 = F.conv2d(img1, window, padding=padd, groups=channel)
    mu2 = F.conv2d(img2, window, padding=padd, groups=channel)

    mu1_sq = mu1.pow(2)
    mu2_sq = mu2.pow(2)
    mu1_mu2 = mu1 * mu2

    sigma1_sq = F.conv2d(img1 * img1, window, padding=padd, groups=channel) - mu1_sq
    sigma2_sq = F.conv2d(img2 * img2, window, padding=padd, groups=channel) - mu2_sq
    sigma12 = F.conv2d(img1 * img2, window, padding=padd, groups=channel) - mu1_mu2

    C1 = (0.01 * L) ** 2
    C2 = (0.03 * L) ** 2

    v1 = 2.0 * sigma12 + C2
    v2 = sigma1_sq + sigma2_sq + C2
    cs = torch.mean(v1 / v2)  # contrast sensitivity

    ssim_map = ((2 * mu1_mu2 + C1) * v1) / ((mu1_sq + mu2_sq + C1) * v2)

    if size_average:
        ret = ssim_map.mean()
    else:
        ret = ssim_map.mean(1).mean(1).mean(1)

    if full:
        return ret, cs
    return ret

class ChannelAttention(nn.Module):
    def __init__(self, in_channel, ratio=16):
        
        super(ChannelAttention, self).__init__()
        self.avg_pool = nn.AdaptiveAvgPool2d(1)
        self.max_pool = nn.AdaptiveMaxPool2d(1)
        self.last_attn = None

        self.fc1 = nn.Conv2d(in_channel, in_channel // ratio, 1, bias=False)
        self.relu1 = nn.ReLU()
        self.fc2 = nn.Conv2d(in_channel // ratio, in_channel, 1, bias=False)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        # print("x.shape",x.shape)
        # avg_out = self.fc2(self.relu1(self.fc1(self.avg_pool(x))))
        # max_out = self.fc2(self.relu1(self.fc1(self.max_pool(x))))
        # out = avg_out + max_out
        # return x*self.sigmoid(out)

        # (2,512,8,8) -> (2,512,1,1)
        avg = self.avg_pool(x)
        # (2,512,1,1) -> (2,512/ratio,1,1)
        avg = self.fc1(avg)
        avg = self.relu1(avg)
        # (2,512/ratio,1,1) -> (2,512,1,1)
        avg_out = self.fc2(avg)

        # (2,512,8,8) -> (2,512,1,1)
        max = self.max_pool(x)
        # (2,512,1,1) -> (2,512/ratio,1,1)
        max = self.fc1(max)
        max = self.relu1(max)
        # (2,512/ratio,1,1) -> (2,512,1,1)
        max_out = self.fc2(max)

        # (2,512,1,1) + (2,512,1,1) -> (2,512,1,1)
        out = avg_out + max_out
        output=self.sigmoid(out)
        if ENABLE_ATTN_ANALYSIS and (not self.training):
            self.last_attn = output.detach()
        else:
            self.last_attn = None
        # print("output.shape",output.shape)
        return x*output

class SpatialAttention(nn.Module):
    def __init__(self, kernel_size=7):
        
        super(SpatialAttention, self).__init__()

        assert kernel_size in (3, 7), 'kernel size must be 3 or 7'
        padding = 3 if kernel_size == 7 else 1

        self.conv1 = nn.Conv2d(2, 1, kernel_size, padding=padding, bias=False)
        self.sigmoid = nn.Sigmoid()
        self.last_attn = None

    def forward(self, x):
        # print('x shape', x.shape)
        # (2,512,8,8) -> (2,1,8,8)
        avg_out = torch.mean(x, dim=1, keepdim=True)
        # (2,512,8,8) -> (2,1,8,8)
        max_out, _ = torch.max(x, dim=1, keepdim=True)
        # (2,1,8,8) + (2,1,8,8) -> (2,2,8,8)
        cat = torch.cat([avg_out, max_out], dim=1)
        # (2,2,8,8) -> (2,1,8,8)
        out = self.conv1(cat)
        output=self.sigmoid(out)
        if ENABLE_ATTN_ANALYSIS and (not self.training):
            self.last_attn = output.detach()
        else:
            self.last_attn = None
        # print("output.shape", output.shape)
        return x*output

# Classes to re-use window
class SSIM(torch.nn.Module):
    def __init__(self, window_size=11, size_average=True, val_range=None):
        super(SSIM, self).__init__()
        self.window_size = window_size
        self.size_average = size_average
        self.val_range = val_range

        # Assume 1 channel for SSIM
        self.channel = 1
        self.window = create_window(window_size)

    def forward(self, img1, img2):
        (_, channel, _, _) = img1.size()

        if channel == self.channel and self.window.dtype == img1.dtype:
            window = self.window
        else:
            window = create_window(self.window_size, channel).to(img1.device).type(img1.dtype)
            self.window = window
            self.channel = channel

        return ssim(img1, img2, window=window, window_size=self.window_size, size_average=self.size_average)
class CNN(nn.Module):
    def __init__(self,ratio,kernel_size):
        super(CNN,self).__init__()
        self.conv = nn.Sequential(
            nn.Conv2d(IMAGE_CHANNELS, 10, 3,padding=1),
            nn.BatchNorm2d(10),
            nn.ReLU(),
            ChannelAttention(10, ratio=ratio),
            SpatialAttention(kernel_size=kernel_size),
            nn.MaxPool2d(2, 2),

            nn.Conv2d(10, 16, 3,padding=1),
            nn.BatchNorm2d(16),
            nn.ReLU(),
            ChannelAttention(16, ratio=ratio),
            SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(16, 21, 3,padding=1),
            nn.BatchNorm2d(21),
            nn.ReLU(),
            ChannelAttention(21, ratio=ratio),
            SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(21, 24, 3,padding=1),
            nn.BatchNorm2d(24),
            nn.ReLU(),
            ChannelAttention(24, ratio=ratio),
            SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(24, 32, 3,padding=1),
            nn.BatchNorm2d(32),
            nn.ReLU(),
            ChannelAttention(32, ratio=ratio),
            SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(32, 36, 3,padding=1),
            nn.BatchNorm2d(36),
            nn.ReLU(),
            ChannelAttention(36, ratio=ratio),
            SpatialAttention(kernel_size=kernel_size),
        )
        self.fc = nn.Sequential(
            nn.Linear(ENCODER_FLAT_DIM, 120),
            torch.nn.BatchNorm1d(120),
            nn.Dropout(0.2),
            nn.ReLU(),
            nn.Linear(120, 84),
            torch.nn.BatchNorm1d(84),
            nn.Dropout(0.2),
            nn.ReLU(),
            nn.Linear(84, 20),
            torch.nn.BatchNorm1d(20),
            nn.Dropout(0.2),
            nn.ReLU(),
            nn.Linear(20, 6),

        )

    def forward(self, x):
        y = self.conv(x)
        output = self.fc(y.view(x.shape[0], -1))
        return output
class ForwardCNN(nn.Module):
    def __init__(self,kernel_size,ratio):
        super( ForwardCNN,self).__init__()
        self.fc = nn.Sequential(
            nn.Linear(6, 20),
            nn.Linear(20, 84),
            torch.nn.BatchNorm1d(84),
            # nn.Dropout(0.2),
            nn.ReLU(),
            nn.Linear(84, 120),
            torch.nn.BatchNorm1d(120),
            # nn.Dropout(0.2),
            nn.ReLU(),
            nn.Linear(120, ENCODER_FLAT_DIM),
            torch.nn.BatchNorm1d(ENCODER_FLAT_DIM),
            nn.Dropout(0.2),
            nn.ReLU(),
        )
        self.conv = nn.Sequential(
            nn.Upsample(size=MODEL_INPUT_SIZE),
            nn.Conv2d(36, 32, 3,padding=1),
            nn.BatchNorm2d(32),
            nn.ReLU(),
            ChannelAttention(32, ratio=ratio),
            SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(32, 24, 3,padding=1),
            nn.BatchNorm2d(24),
            nn.ReLU(),
            ChannelAttention(24, ratio=ratio),
            SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(24, 21, 3,padding=1),
            nn.BatchNorm2d(21),
            nn.ReLU(),
            ChannelAttention(21, ratio=ratio),
            SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(21, 16, 3,padding=1),
            nn.BatchNorm2d(16),
            nn.ReLU(),
            ChannelAttention(16, ratio=ratio),
            SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(16, 10, 3,padding=1),
            nn.BatchNorm2d(10),
            nn.ReLU(),
            ChannelAttention(10, ratio=ratio),
            SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(10, IMAGE_CHANNELS, 3,padding=1),
            nn.BatchNorm2d(IMAGE_CHANNELS),
            nn.ReLU(),
            ChannelAttention(IMAGE_CHANNELS, ratio=1),
            SpatialAttention(kernel_size=kernel_size),
        )

    def forward(self, x):
        y = self.fc(x.view(x.shape[0], -1))
        y=torch.unsqueeze(y,dim=2)
        y=torch.unsqueeze(y,dim=3)
        y=torch.reshape(y,(-1, ENCODER_CHANNELS, ENCODER_H, ENCODER_W))
        output= self.conv(y)
        return output


def _model_attn_counts(model):
    spatial = 0
    channel = 0
    for m in model.modules():
        name = m.__class__.__name__
        if name == "SpatialAttention":
            spatial += 1
        elif name == "ChannelAttention":
            channel += 1
    return spatial, channel


def _extract_state_dict_from_checkpoint(obj):
    if isinstance(obj, dict):
        for k in ["state_dict", "model_state_dict", "net", "model"]:
            if k in obj and isinstance(obj[k], dict):
                return obj[k]
        if len(obj) > 0 and all(isinstance(v, torch.Tensor) for v in obj.values()):
            return obj
    if isinstance(obj, nn.Module):
        return obj.state_dict()
    return None


def _strip_module_prefix(state_dict):
    if not isinstance(state_dict, dict) or len(state_dict) == 0:
        return state_dict
    keys = list(state_dict.keys())
    if all(k.startswith("module.") for k in keys):
        return {k[len("module."):]: v for k, v in state_dict.items()}
    return state_dict


def load_inverse_model_with_cbam_guard(ckpt_path, device, kernel_size=3, ratio=2):
    obj = torch.load(str(ckpt_path), map_location=device)
    if isinstance(obj, nn.Module):
        candidate = obj.to(device)
        s_cnt, c_cnt = _model_attn_counts(candidate)
        if s_cnt > 0:
            return candidate, {"mode": "module_direct", "spatial": s_cnt, "channel": c_cnt}

    # Fallback: load checkpoint weights into current CNN class to ensure CBAM modules exist.
    state_dict = _extract_state_dict_from_checkpoint(obj)
    if state_dict is not None:
        state_dict = _strip_module_prefix(state_dict)
        rebuilt = CNN(kernel_size=kernel_size, ratio=ratio).to(device)
        missing, unexpected = rebuilt.load_state_dict(state_dict, strict=False)
        s_cnt, c_cnt = _model_attn_counts(rebuilt)
        return rebuilt, {
            "mode": "state_dict_into_current_cnn",
            "spatial": s_cnt,
            "channel": c_cnt,
            "missing_keys": len(missing),
            "unexpected_keys": len(unexpected)
        }

    if isinstance(obj, nn.Module):
        candidate = obj.to(device)
        s_cnt, c_cnt = _model_attn_counts(candidate)
        return candidate, {"mode": "module_no_cbam", "spatial": s_cnt, "channel": c_cnt}

    raise RuntimeError(f"Unsupported checkpoint format for inverse model: {type(obj)}")

# Run mode: "train" trains then tests; "test" loads checkpoints only
RUN_STAGE = "test"  # "train" or "test"
MODEL_TAG = "cnn-1"
TRAIN_BATCH_SIZE = 16
TEST_BATCH_SIZE = 8
RESUME_TRAIN = True


# Batch-wise test inference to avoid OOM
def run_test_in_batches(model_inv, model_fwd, x_all, y_all=None, batch_size=TEST_BATCH_SIZE):
    if isinstance(x_all, Data.Dataset):
        test_set_local = x_all
    else:
        if y_all is None:
            raise ValueError("y_all must be provided when x_all is a tensor.")
        test_set_local = Data.TensorDataset(x_all, y_all)

    if len(test_set_local) == 0:
        raise RuntimeError("Empty test set. Cannot run testing.")
    test_loader_local = Data.DataLoader(
        dataset=test_set_local,
        batch_size=batch_size,
        num_workers=0,
        shuffle=False
    )

    model_inv.eval()
    model_fwd.eval()
    out_list, label_list, stru_list, img_list = [], [], [], []
    ssim_sum = 0.0
    n_sum = 0
    with torch.no_grad():
        for xb, yb in test_loader_local:
            xb = xb.to(device)
            yb = yb.to(device)
            outb = model_inv(xb)
            strub = model_fwd(outb)
            bs = xb.size(0)
            ssim_b = float(ssim(strub.detach(), xb.detach(), size_average=True).item())
            ssim_sum += ssim_b * bs
            n_sum += bs
            out_list.append(outb.cpu())
            label_list.append(yb.cpu())
            stru_list.append(strub.cpu())
            img_list.append(xb.cpu())

    if n_sum == 0:
        raise RuntimeError("Empty test loader. Cannot compute results.")

    return (
        torch.cat(img_list, dim=0),
        torch.cat(label_list, dim=0),
        torch.cat(out_list, dim=0),
        torch.cat(stru_list, dim=0),
        ssim_sum / n_sum
    )


# -----------------------------
# Attention extraction utilities
# -----------------------------
def _is_spatial_attn_module(m):
    if isinstance(m, SpatialAttention) or m.__class__.__name__ == "SpatialAttention":
        return True
    # Heuristic for cross-script / legacy checkpoints:
    # Spatial attention usually has conv1(2->1) + sigmoid and outputs x * attn.
    conv1 = getattr(m, "conv1", None)
    sig = getattr(m, "sigmoid", None)
    if isinstance(conv1, nn.Conv2d) and isinstance(sig, nn.Sigmoid):
        if conv1.in_channels == 2 and conv1.out_channels == 1:
            return True
    return False


def _is_channel_attn_module(m):
    return isinstance(m, ChannelAttention) or m.__class__.__name__ == "ChannelAttention"


def clear_attention_cache(model):
    for m in model.modules():
        if _is_spatial_attn_module(m) or _is_channel_attn_module(m):
            setattr(m, "last_attn", None)


def collect_spatial_attn(model):
    maps = []
    for m in model.modules():
        if _is_spatial_attn_module(m):
            attn = getattr(m, "last_attn", None)
            if isinstance(attn, torch.Tensor):
                maps.append(attn)
    return maps


def collect_channel_attn(model):
    maps = []
    for m in model.modules():
        if _is_channel_attn_module(m):
            attn = getattr(m, "last_attn", None)
            if isinstance(attn, torch.Tensor):
                maps.append(attn)
    return maps


def _forward_collect_attn_with_hooks(model, x):
    spatial_out = []
    channel_out = []
    handles = []

    def _make_spatial_hook(store):
        def _hook(module, inp, out):
            if len(inp) == 0 or not isinstance(inp[0], torch.Tensor):
                return
            x_in = inp[0].detach()
            attn = None
            conv1 = getattr(module, "conv1", None)
            sig = getattr(module, "sigmoid", None)
            # Preferred: reconstruct by the module's own conv1+sigmoid path.
            if isinstance(conv1, nn.Conv2d) and isinstance(sig, nn.Sigmoid):
                avg_out = torch.mean(x_in, dim=1, keepdim=True)
                max_out, _ = torch.max(x_in, dim=1, keepdim=True)
                cat = torch.cat([avg_out, max_out], dim=1)
                attn = sig(conv1(cat)).detach()
            # Fallback: infer from output/input ratio.
            elif isinstance(out, torch.Tensor) and out.shape == x_in.shape:
                ratio = out.detach() / (x_in + 1e-8)
                attn = torch.clamp(torch.mean(ratio, dim=1, keepdim=True), 0.0, 1.0)
            if isinstance(attn, torch.Tensor):
                store.append(attn)
        return _hook

    def _make_channel_hook(store):
        def _hook(_module, _inp, out):
            if isinstance(out, torch.Tensor):
                store.append(out.detach())
        return _hook

    for m in model.modules():
        if _is_spatial_attn_module(m):
            handles.append(m.register_forward_hook(_make_spatial_hook(spatial_out)))
        elif _is_channel_attn_module(m):
            s = getattr(m, "sigmoid", None)
            if isinstance(s, nn.Sigmoid):
                handles.append(s.register_forward_hook(_make_channel_hook(channel_out)))

    try:
        _ = model(x)
    finally:
        for h in handles:
            h.remove()

    return spatial_out, channel_out


def forward_collect_attn(model, x, prefer_hook=False):
    clear_attention_cache(model)
    if not prefer_hook:
        _ = model(x)
        s_list = collect_spatial_attn(model)
        c_list = collect_channel_attn(model)
        if len(s_list) > 0:
            return s_list, c_list, False
    # Fallback: for old checkpoints/classes where cache is unavailable.
    s_list, c_list = _forward_collect_attn_with_hooks(model, x)
    return s_list, c_list, True


def _normalize_b1hw(x, eps=1e-8):
    b = x.shape[0]
    x_flat = x.view(b, -1)
    x_min = x_flat.min(dim=1, keepdim=True)[0].view(b, 1, 1, 1)
    x_max = x_flat.max(dim=1, keepdim=True)[0].view(b, 1, 1, 1)
    return (x - x_min) / (x_max - x_min + eps)


def aggregate_spatial_attn(attn_list, target_hw, method="mean"):
    if len(attn_list) == 0:
        raise RuntimeError("No spatial attention map found. Ensure model_inverse has run in eval mode.")
    resized = [
        F.interpolate(a, size=target_hw, mode="bilinear", align_corners=False)
        for a in attn_list
    ]
    stack = torch.stack(resized, dim=0)  # (L, B, 1, H, W)
    if method == "weighted":
        # Shallow layers are assigned larger weights for fine-grained localization.
        w = torch.linspace(float(len(resized)), 1.0, steps=len(resized), device=stack.device, dtype=stack.dtype)
        w = w / w.sum()
        agg = (stack * w.view(-1, 1, 1, 1, 1)).sum(dim=0)
    elif method == "max":
        # Preserve the strongest response among layers. This is often more robust than
        # mean aggregation when some layers are close to spatially uniform.
        agg = stack.max(dim=0).values
    else:
        agg = stack.mean(dim=0)
    return _normalize_b1hw(agg)


# ------------------------
# Visualization utilities
# ------------------------
def _img_chw_to_gray_np(img_chw):
    if isinstance(img_chw, torch.Tensor):
        x = img_chw.detach().float().cpu()
    else:
        x = torch.tensor(img_chw, dtype=torch.float32)
    if x.ndim == 3:
        if x.shape[0] >= 3:
            g = x[:3].mean(dim=0)
        else:
            g = x[0]
    elif x.ndim == 2:
        g = x
    else:
        raise ValueError(f"Unexpected image shape: {tuple(x.shape)}")
    g = torch.clamp(g, 0.0, 1.0).numpy()
    return g


def _attn_1hw_to_np(attn_1hw):
    if isinstance(attn_1hw, torch.Tensor):
        a = attn_1hw.detach().float().cpu().numpy()
    else:
        a = np.asarray(attn_1hw, dtype=np.float32)
    if a.ndim == 3:
        a = a[0]
    a = np.asarray(a, dtype=np.float32)
    a = (a - a.min()) / (a.max() - a.min() + 1e-8)
    return a


def save_attention_overlay(img_chw, attn_1hw, sample_tag, overlay_dir=ATTN_OVERLAY_DIR, raw_dir=ATTN_RAW_DIR):
    gray = _img_chw_to_gray_np(img_chw)
    attn = _attn_1hw_to_np(attn_1hw)
    heat_rgb = plt.cm.jet(attn)[..., :3]
    gray_rgb = np.stack([gray, gray, gray], axis=-1)
    overlay = np.clip(0.60 * gray_rgb + 0.40 * heat_rgb, 0.0, 1.0)

    old_family = plt.rcParams.get("font.family", None)
    plt.rcParams["font.family"] = "Times New Roman"
    fig, axes = plt.subplots(1, 3, figsize=(12, 4), dpi=300)
    panels = [
        ("Input Image", gray, "gray"),
        ("Aggregated Attention", attn, "jet"),
        ("Overlay", overlay, None),
    ]
    for ax, (title, data, cmap) in zip(axes, panels):
        if cmap is None:
            ax.imshow(data)
        else:
            ax.imshow(data, cmap=cmap)
        ax.set_title(title, fontsize=11, fontweight="bold")
        ax.set_xlabel("X", fontsize=9, fontweight="bold")
        ax.set_ylabel("Y", fontsize=9, fontweight="bold")
        ax.set_xticks([])
        ax.set_yticks([])
    fig.tight_layout()
    fig.savefig(str(overlay_dir / f"{sample_tag}.png"), dpi=300, bbox_inches="tight")
    plt.close(fig)
    if old_family is not None:
        plt.rcParams["font.family"] = old_family
    if ATTN_SAVE_RAW_MAPS:
        np.save(str(raw_dir / f"{sample_tag}_agg.npy"), attn.astype(np.float32))


def save_attention_grid(img_chw, per_layer_attn, sample_tag, per_layer_dir=ATTN_PER_LAYER_DIR):
    if len(per_layer_attn) == 0:
        return
    gray = _img_chw_to_gray_np(img_chw)
    maps = [_attn_1hw_to_np(m) for m in per_layer_attn]
    n_panels = 1 + len(maps)
    ncols = 3
    nrows = int(np.ceil(n_panels / float(ncols)))

    old_family = plt.rcParams.get("font.family", None)
    plt.rcParams["font.family"] = "Times New Roman"
    fig, axes = plt.subplots(nrows, ncols, figsize=(4 * ncols, 3 * nrows), dpi=300)
    axes = np.array(axes).reshape(-1)

    axes[0].imshow(gray, cmap="gray")
    axes[0].set_title("Input Image", fontsize=11, fontweight="bold")
    axes[0].set_xlabel("X", fontsize=9, fontweight="bold")
    axes[0].set_ylabel("Y", fontsize=9, fontweight="bold")
    axes[0].set_xticks([])
    axes[0].set_yticks([])

    for i, attn in enumerate(maps):
        ax = axes[i + 1]
        ax.imshow(attn, cmap="jet")
        ax.set_title(f"SpatialAttn L{i+1}", fontsize=11, fontweight="bold")
        ax.set_xlabel("X", fontsize=9, fontweight="bold")
        ax.set_ylabel("Y", fontsize=9, fontweight="bold")
        ax.set_xticks([])
        ax.set_yticks([])

    for j in range(n_panels, len(axes)):
        axes[j].axis("off")
    fig.tight_layout()
    fig.savefig(str(per_layer_dir / f"{sample_tag}_per_layer.png"), dpi=300, bbox_inches="tight")
    plt.close(fig)
    if old_family is not None:
        plt.rcParams["font.family"] = old_family


# -----------------------------------
# Quantitative interpretability metrics
# -----------------------------------
def sobel_edge_map(img_b1hw):
    kx = torch.tensor([[-1.0, 0.0, 1.0],
                       [-2.0, 0.0, 2.0],
                       [-1.0, 0.0, 1.0]], device=img_b1hw.device, dtype=img_b1hw.dtype).view(1, 1, 3, 3)
    ky = torch.tensor([[-1.0, -2.0, -1.0],
                       [0.0, 0.0, 0.0],
                       [1.0, 2.0, 1.0]], device=img_b1hw.device, dtype=img_b1hw.dtype).view(1, 1, 3, 3)
    gx = F.conv2d(img_b1hw, kx, padding=1)
    gy = F.conv2d(img_b1hw, ky, padding=1)
    return torch.sqrt(gx * gx + gy * gy + 1e-12)


def _valid_inner_mask_like(x_b1hw, border_ignore=ATTN_BORDER_IGNORE):
    b, _, h, w = x_b1hw.shape
    mask = torch.ones((b, 1, h, w), device=x_b1hw.device, dtype=torch.bool)
    if border_ignore > 0 and h > 2 * border_ignore and w > 2 * border_ignore:
        mask[:, :, :border_ignore, :] = False
        mask[:, :, -border_ignore:, :] = False
        mask[:, :, :, :border_ignore] = False
        mask[:, :, :, -border_ignore:] = False
    return mask


def _dilate_binary_mask(mask_b1hw, radius):
    if radius <= 0:
        return mask_b1hw.bool()
    k = int(radius) * 2 + 1
    pooled = F.max_pool2d(mask_b1hw.float(), kernel_size=k, stride=1, padding=radius)
    return pooled > 0.5


def _build_foreground_mask(gray_b1hw, thresh=ATTN_FOREGROUND_THRESH, border_ignore=ATTN_BORDER_IGNORE):
    # Images are near-binary (bright structures on dark background). Use a conservative
    # global threshold, then remove border regions to avoid padding artifacts.
    valid = _valid_inner_mask_like(gray_b1hw, border_ignore=border_ignore)
    fg = gray_b1hw > float(thresh)
    fg = fg & valid
    return fg, valid


def compute_attention_region_stats(
    img_bchw,
    attn_b1hw,
    topk=ATTN_EDGE_TOPK,
    border_ignore=ATTN_BORDER_IGNORE,
    fg_thresh=ATTN_FOREGROUND_THRESH,
    edge_band_width=ATTN_EDGE_BAND_WIDTH,
):
    gray = img_bchw.mean(dim=1, keepdim=True)
    fg_mask, valid_mask = _build_foreground_mask(gray, thresh=fg_thresh, border_ignore=border_ignore)
    dilated_fg = _dilate_binary_mask(fg_mask, radius=edge_band_width)
    background_mask = valid_mask & (~dilated_fg)

    edge = sobel_edge_map(gray)
    b = edge.shape[0]
    flat_edge = edge.view(b, -1)
    edge_topk_mask = torch.zeros_like(edge, dtype=torch.bool)
    k = max(1, int(flat_edge.shape[1] * float(topk)))
    for i in range(b):
        valid_flat = valid_mask[i].view(-1)
        edge_i = flat_edge[i].clone()
        edge_i[~valid_flat] = -1e9
        _, idx = torch.topk(edge_i, k=k, largest=True)
        mask_i = torch.zeros_like(edge_i, dtype=torch.bool)
        mask_i[idx] = True
        edge_topk_mask[i] = mask_i.view_as(edge_topk_mask[i])

    edge_band_mask = _dilate_binary_mask(edge_topk_mask & fg_mask, radius=edge_band_width)
    edge_band_mask = edge_band_mask & valid_mask
    edge_band_mask = edge_band_mask & (~background_mask)
    far_background_mask = background_mask

    flat_attn = attn_b1hw.view(b, -1)
    flat_fg = fg_mask.view(b, -1)
    flat_bg = background_mask.view(b, -1)
    flat_edge_band = edge_band_mask.view(b, -1)
    flat_far_bg = far_background_mask.view(b, -1)

    fg_mean_list = []
    bg_mean_list = []
    fg_ratio_list = []
    edge_band_mean_list = []
    far_bg_mean_list = []
    edge_band_ratio_list = []
    edge_strength = []
    fg_fraction = []

    for i in range(b):
        attn_i = flat_attn[i]
        fg_i = flat_fg[i]
        bg_i = flat_bg[i]
        eb_i = flat_edge_band[i]
        fbg_i = flat_far_bg[i]

        fg_mean = float(attn_i[fg_i].mean().item()) if fg_i.any() else 0.0
        bg_mean = float(attn_i[bg_i].mean().item()) if bg_i.any() else 0.0
        eb_mean = float(attn_i[eb_i].mean().item()) if eb_i.any() else 0.0
        fbg_mean = float(attn_i[fbg_i].mean().item()) if fbg_i.any() else bg_mean

        fg_ratio = fg_mean / (bg_mean + 1e-8)
        eb_ratio = eb_mean / (fbg_mean + 1e-8)

        fg_mean_list.append(fg_mean)
        bg_mean_list.append(bg_mean)
        fg_ratio_list.append(fg_ratio)
        edge_band_mean_list.append(eb_mean)
        far_bg_mean_list.append(fbg_mean)
        edge_band_ratio_list.append(eb_ratio)
        edge_strength.append(float((edge[i] * valid_mask[i].float()).sum().item() / (valid_mask[i].float().sum().item() + 1e-8)))
        fg_fraction.append(float(fg_i.float().mean().item()))

    return {
        "foreground_attn_mean": np.array(fg_mean_list, dtype=np.float32),
        "background_attn_mean": np.array(bg_mean_list, dtype=np.float32),
        "foreground_attention_ratio": np.array(fg_ratio_list, dtype=np.float32),
        "edge_band_attn_mean": np.array(edge_band_mean_list, dtype=np.float32),
        "far_background_attn_mean": np.array(far_bg_mean_list, dtype=np.float32),
        "edge_band_attention_ratio": np.array(edge_band_ratio_list, dtype=np.float32),
        "edge_strength": np.array(edge_strength, dtype=np.float32),
        "foreground_fraction": np.array(fg_fraction, dtype=np.float32),
    }


def compute_texture_attention_correlation(img_bchw, attn_b1hw, border_ignore=ATTN_BORDER_IGNORE):
    gray = img_bchw.mean(dim=1, keepdim=True)
    valid = _valid_inner_mask_like(gray, border_ignore=border_ignore)
    lap = torch.tensor([[0.0, 1.0, 0.0],
                        [1.0, -4.0, 1.0],
                        [0.0, 1.0, 0.0]], device=gray.device, dtype=gray.dtype).view(1, 1, 3, 3)
    texture = torch.abs(F.conv2d(gray, lap, padding=1))
    b = texture.shape[0]
    corr_list = []
    for i in range(b):
        mask = valid[i].reshape(-1)
        t = texture[i].reshape(-1)[mask]
        a = attn_b1hw[i].reshape(-1)[mask]
        t = t - t.mean()
        a = a - a.mean()
        denom = torch.sqrt(torch.sum(t * t) * torch.sum(a * a) + 1e-12)
        corr = torch.sum(t * a) / (denom + 1e-12)
        corr_list.append(float(corr.item()))
    return np.array(corr_list, dtype=np.float32)



def _percentile_rank(series):
    return series.rank(pct=True, method="average")


def _foreground_fraction_score(x, lo=ATTN_REP_MIN_FOREGROUND, hi=ATTN_REP_MAX_FOREGROUND):
    # Prefer moderate foreground occupancy: too tiny gives unstable ratios,
    # too large tends to blur "attention vs background" contrast.
    if lo <= x <= hi:
        return 1.0
    if x < lo:
        return max(0.0, float(x / max(lo, 1e-8)))
    # Linearly decay after hi and clamp to zero near 0.85 occupancy.
    return max(0.0, float(1.0 - (x - hi) / max(0.85 - hi, 1e-8)))


def add_representative_scores(sample_df):
    df = sample_df.copy()
    # Raw ratios can explode when the background denominator is near zero.
    # Use log-ratios and percentile ranks for robust ranking/selection.
    df["fg_ratio_log1p"] = np.log1p(np.clip(df["foreground_attention_ratio"].to_numpy(dtype=np.float64), 0.0, None))
    df["edge_ratio_log1p"] = np.log1p(np.clip(df["edge_band_attention_ratio"].to_numpy(dtype=np.float64), 0.0, None))

    df["fg_ratio_rank"] = _percentile_rank(df["fg_ratio_log1p"])
    df["edge_ratio_rank"] = _percentile_rank(df["edge_ratio_log1p"])
    df["texture_rank"] = _percentile_rank(df["texture_attn_corr"])
    df["edge_strength_rank"] = _percentile_rank(df["edge_strength"])
    df["foreground_fraction_score"] = df["foreground_fraction"].apply(_foreground_fraction_score)

    # Global persuasive score for qualitative figures:
    # structure-focused ratio + edge-band ratio + texture agreement + complexity.
    df["representative_score_raw"] = (
        0.30 * df["fg_ratio_rank"] +
        0.30 * df["edge_ratio_rank"] +
        0.25 * df["texture_rank"] +
        0.15 * df["edge_strength_rank"]
    )
    df["representative_score"] = df["representative_score_raw"] * df["foreground_fraction_score"]

    # Mark samples that are stable enough for figure selection.
    df["is_stable_candidate"] = (
        (df["foreground_fraction"] >= ATTN_REP_MIN_FOREGROUND) &
        (df["foreground_fraction"] <= ATTN_REP_MAX_FOREGROUND) &
        (df["background_attn_mean"] > ATTN_REP_MIN_BG_ATTENTION) &
        (df["far_background_attn_mean"] > ATTN_REP_MIN_BG_ATTENTION)
    )

    # Complexity group for diverse qualitative examples.
    valid = df["edge_strength"].to_numpy(dtype=np.float64)
    q1, q2 = np.quantile(valid, [1/3, 2/3])
    groups = []
    for v in valid:
        if v <= q1:
            groups.append("low")
        elif v <= q2:
            groups.append("mid")
        else:
            groups.append("high")
    df["complexity_group"] = groups

    return df


def select_representative_samples(sample_df, manual_ids=None, n_each=2):
    n_total = len(sample_df)
    if n_total == 0:
        return [], pd.DataFrame()

    df = add_representative_scores(sample_df)

    if manual_ids is not None and len(manual_ids) > 0:
        chosen = [int(i) for i in manual_ids if 0 <= int(i) < n_total]
        rep_df = df[df["sample_idx"].isin(chosen)].copy()
        rep_df["selection_reason"] = "manual"
        rep_df = rep_df.sort_values(["sample_idx"]).reset_index(drop=True)
        return sorted(list(set(chosen))), rep_df

    candidate_df = df[df["is_stable_candidate"]].copy()
    if len(candidate_df) == 0:
        candidate_df = df.copy()

    candidate_df = candidate_df.sort_values(
        ["representative_score", "texture_attn_corr", "edge_strength"],
        ascending=[False, False, False]
    ).reset_index(drop=True)

    picked_rows = []
    picked_ids = set()

    # First pass: choose diverse samples from low/mid/high complexity bins.
    for group in ["low", "mid", "high"]:
        sub = candidate_df[candidate_df["complexity_group"] == group]
        taken = 0
        for _, row in sub.iterrows():
            sid = int(row["sample_idx"])
            if sid in picked_ids:
                continue
            picked_rows.append((sid, group, "diverse-top"))
            picked_ids.add(sid)
            taken += 1
            if taken >= n_each:
                break

    # Second pass: if some bins were sparse, fill by global score.
    target_n = max(1, 3 * n_each)
    if len(picked_rows) < target_n:
        for _, row in candidate_df.iterrows():
            sid = int(row["sample_idx"])
            if sid in picked_ids:
                continue
            picked_rows.append((sid, str(row["complexity_group"]), "global-fill"))
            picked_ids.add(sid)
            if len(picked_rows) >= target_n:
                break

    rep_df = pd.DataFrame(picked_rows, columns=["sample_idx", "selected_group", "selection_reason"])
    rep_df = rep_df.merge(df, on="sample_idx", how="left")
    rep_df = rep_df.sort_values(["selected_group", "representative_score"], ascending=[True, False]).reset_index(drop=True)
    return rep_df["sample_idx"].astype(int).tolist(), rep_df


def run_attention_interpretability_analysis(model_inv, x_all, batch_size=TEST_BATCH_SIZE):
    if len(x_all) == 0:
        raise RuntimeError("Empty test set for attention analysis.")

    model_inv.eval()
    recognized_sa = sum(1 for m in model_inv.modules() if _is_spatial_attn_module(m))
    print(f"[AttentionAnalysis] recognized SpatialAttention-like modules: {recognized_sa}")
    spatial_cnt, channel_cnt = _model_attn_counts(model_inv)
    if spatial_cnt == 0:
        uniq = []
        for m in model_inv.modules():
            n = m.__class__.__name__
            if n not in uniq:
                uniq.append(n)
        raise RuntimeError(
            "Loaded inverse model has no SpatialAttention modules "
            f"(spatial={spatial_cnt}, channel={channel_cnt}). "
            "Likely checkpoint/model mismatch. "
            f"Model class={model_inv.__class__.__name__}, unique modules(head)={uniq[:20]}"
        )

    loader = Data.DataLoader(
        dataset=Data.TensorDataset(x_all),
        batch_size=batch_size,
        num_workers=0,
        shuffle=False
    )

    rows = []
    spatial_hw_shapes = None
    channel_shapes = None
    sample_base = 0
    used_hook_fallback = False
    with torch.no_grad():
        for (xb,) in loader:
            xb = xb.to(device)
            s_list, c_list, used_hook_fallback = forward_collect_attn(
                model_inv, xb, prefer_hook=used_hook_fallback
            )
            if len(s_list) == 0:
                raise RuntimeError(
                    "Spatial attention extraction failed (cache and hook fallback both empty). "
                    "Check whether inverse model contains SpatialAttention modules."
                )
            if spatial_hw_shapes is None:
                spatial_hw_shapes = [tuple(m.shape[-2:]) for m in s_list]
            if channel_shapes is None:
                channel_shapes = [int(m.shape[1]) for m in c_list]

            agg = aggregate_spatial_attn(
                s_list,
                target_hw=(xb.shape[2], xb.shape[3]),
                method=ATTN_AGG_METHOD
            )
            region_res = compute_attention_region_stats(
                xb, agg, topk=ATTN_EDGE_TOPK, border_ignore=ATTN_BORDER_IGNORE,
                fg_thresh=ATTN_FOREGROUND_THRESH, edge_band_width=ATTN_EDGE_BAND_WIDTH
            )
            tex_corr = compute_texture_attention_correlation(xb, agg, border_ignore=ATTN_BORDER_IGNORE)

            bs = xb.shape[0]
            for i in range(bs):
                rows.append({
                    "sample_idx": int(sample_base + i),
                    "foreground_attn_mean": float(region_res["foreground_attn_mean"][i]),
                    "background_attn_mean": float(region_res["background_attn_mean"][i]),
                    "foreground_attention_ratio": float(region_res["foreground_attention_ratio"][i]),
                    "edge_band_attn_mean": float(region_res["edge_band_attn_mean"][i]),
                    "far_background_attn_mean": float(region_res["far_background_attn_mean"][i]),
                    "edge_band_attention_ratio": float(region_res["edge_band_attention_ratio"][i]),
                    "foreground_fraction": float(region_res["foreground_fraction"][i]),
                    "edge_strength": float(region_res["edge_strength"][i]),
                    "texture_attn_corr": float(tex_corr[i]),
                })
            sample_base += bs

    sample_df = pd.DataFrame(rows)
    sample_df = add_representative_scores(sample_df)
    sample_df.to_csv(str(ATTN_ROOT / "attention_sample_summary.csv"), index=False, encoding="utf-8-sig")

    fg_mean = float(sample_df["foreground_attention_ratio"].mean())
    fg_std = float(sample_df["foreground_attention_ratio"].std(ddof=0))
    fg_median = float(sample_df["foreground_attention_ratio"].median())
    fg_q25 = float(sample_df["foreground_attention_ratio"].quantile(0.25))
    fg_q75 = float(sample_df["foreground_attention_ratio"].quantile(0.75))
    fg_stats_df = pd.DataFrame([{
        "n_samples": int(len(sample_df)),
        "foreground_threshold": float(ATTN_FOREGROUND_THRESH),
        "border_ignore": int(ATTN_BORDER_IGNORE),
        "foreground_attention_ratio_mean": fg_mean,
        "foreground_attention_ratio_std": fg_std,
        "foreground_attention_ratio_median": fg_median
    }])
    fg_stats_df.to_csv(str(ATTN_ROOT / "foreground_attention_stats.csv"), index=False, encoding="utf-8-sig")

    edge_mean = float(sample_df["edge_band_attention_ratio"].mean())
    edge_std = float(sample_df["edge_band_attention_ratio"].std(ddof=0))
    edge_median = float(sample_df["edge_band_attention_ratio"].median())
    edge_q25 = float(sample_df["edge_band_attention_ratio"].quantile(0.25))
    edge_q75 = float(sample_df["edge_band_attention_ratio"].quantile(0.75))
    edge_stats_df = pd.DataFrame([{
        "n_samples": int(len(sample_df)),
        "topk": float(ATTN_EDGE_TOPK),
        "edge_band_width": int(ATTN_EDGE_BAND_WIDTH),
        "border_ignore": int(ATTN_BORDER_IGNORE),
        "edge_band_attention_ratio_mean": edge_mean,
        "edge_band_attention_ratio_std": edge_std,
        "edge_band_attention_ratio_median": edge_median
    }])
    edge_stats_df.to_csv(str(ATTN_ROOT / "edge_attention_stats.csv"), index=False, encoding="utf-8-sig")

    tex_mean = float(sample_df["texture_attn_corr"].mean())
    tex_std = float(sample_df["texture_attn_corr"].std(ddof=0))
    tex_median = float(sample_df["texture_attn_corr"].median())
    tex_q25 = float(sample_df["texture_attn_corr"].quantile(0.25))
    tex_q75 = float(sample_df["texture_attn_corr"].quantile(0.75))
    tex_stats_df = pd.DataFrame([{
        "n_samples": int(len(sample_df)),
        "border_ignore": int(ATTN_BORDER_IGNORE),
        "texture_corr_mean": tex_mean,
        "texture_corr_std": tex_std,
        "texture_corr_median": tex_median
    }])
    tex_stats_df.to_csv(str(ATTN_ROOT / "texture_attention_stats.csv"), index=False, encoding="utf-8-sig")

    robust_summary_df = pd.DataFrame([{
        "n_samples": int(len(sample_df)),
        "foreground_ratio_median": fg_median,
        "foreground_ratio_q25": fg_q25,
        "foreground_ratio_q75": fg_q75,
        "edge_band_ratio_median": edge_median,
        "edge_band_ratio_q25": edge_q25,
        "edge_band_ratio_q75": edge_q75,
        "texture_corr_mean": tex_mean,
        "texture_corr_median": tex_median,
        "texture_corr_q25": tex_q25,
        "texture_corr_q75": tex_q75
    }])
    robust_summary_df.to_csv(str(ATTN_ROOT / "attention_robust_summary.csv"), index=False, encoding="utf-8-sig")

    selected_ids, selected_df = select_representative_samples(
        sample_df,
        manual_ids=ATTN_VIS_SAMPLE_IDS,
        n_each=ATTN_AUTO_SAMPLES_PER_GROUP
    )
    ranked_df = sample_df[sample_df["is_stable_candidate"]].copy()
    if len(ranked_df) == 0:
        ranked_df = sample_df.copy()
    ranked_df = ranked_df.sort_values(
        ["representative_score", "texture_attn_corr", "edge_strength"],
        ascending=[False, False, False]
    ).reset_index(drop=True)
    ranked_df["rank"] = np.arange(1, len(ranked_df) + 1)
    ranked_df.head(ATTN_REP_TOPN_EXPORT).to_csv(
        str(ATTN_ROOT / "representative_samples_ranked.csv"),
        index=False, encoding="utf-8-sig"
    )
    selected_df.to_csv(str(ATTN_ROOT / "representative_samples_selected.csv"), index=False, encoding="utf-8-sig")
    selected_meta = {}
    if len(selected_df) > 0:
        for rank_i, row in selected_df.reset_index(drop=True).iterrows():
            selected_meta[int(row["sample_idx"])] = {
                "rank": int(rank_i + 1),
                "group": str(row.get("selected_group", row.get("complexity_group", "rep"))),
                "score": float(row.get("representative_score", 0.0)),
            }

    for sid in selected_ids:
        xb = x_all[sid:sid+1].to(device)
        with torch.no_grad():
            s_list, _, used_hook_fallback = forward_collect_attn(
                model_inv, xb, prefer_hook=used_hook_fallback
            )
            agg = aggregate_spatial_attn(
                s_list,
                target_hw=(xb.shape[2], xb.shape[3]),
                method=ATTN_AGG_METHOD
            )

        meta = selected_meta.get(int(sid), {"rank": 0, "group": "rep", "score": 0.0})
        sample_tag = f"rank{meta['rank']:02d}_{meta['group']}_sample_{sid:04d}"
        save_attention_overlay(xb[0], agg[0], sample_tag, overlay_dir=ATTN_OVERLAY_DIR, raw_dir=ATTN_RAW_DIR)

        if ATTN_SAVE_PER_LAYER:
            per_layer = []
            for m in s_list:
                up = F.interpolate(m, size=(xb.shape[2], xb.shape[3]), mode="bilinear", align_corners=False)
                per_layer.append(_normalize_b1hw(up)[0])
            save_attention_grid(xb[0], per_layer, sample_tag, per_layer_dir=ATTN_PER_LAYER_DIR)

        if ATTN_SAVE_RAW_MAPS:
            per_layer_np = []
            for m in s_list:
                up = F.interpolate(m, size=(xb.shape[2], xb.shape[3]), mode="bilinear", align_corners=False)
                per_layer_np.append(_normalize_b1hw(up)[0, 0].detach().cpu().numpy().astype(np.float32))
            np.save(str(ATTN_RAW_DIR / f"{sample_tag}_per_layer.npy"), np.stack(per_layer_np, axis=0))

    with open(str(ATTN_ROOT / "attention_summary.txt"), "w", encoding="utf-8") as f:
        f.write("Attention Interpretability Summary\n")
        f.write(f"Samples analyzed: {len(sample_df)}\n")
        f.write(f"Spatial attention layer count: {0 if spatial_hw_shapes is None else len(spatial_hw_shapes)}\n")
        f.write(f"Spatial layer sizes (H, W): {spatial_hw_shapes}\n")
        f.write(f"Channel attention channel sizes: {channel_shapes}\n")
        f.write(f"Aggregation method: {ATTN_AGG_METHOD}\n")
        f.write(f"Hook fallback used: {used_hook_fallback}\n")
        f.write(f"Foreground threshold: {ATTN_FOREGROUND_THRESH}\n")
        f.write(f"Border ignore width: {ATTN_BORDER_IGNORE}\n")
        f.write(f"Edge top-k ratio: {ATTN_EDGE_TOPK}\n")
        f.write(f"Edge band width: {ATTN_EDGE_BAND_WIDTH}\n")
        f.write(f"Foreground attention ratio mean/std/median: {fg_mean:.6f} / {fg_std:.6f} / {fg_median:.6f}\n")
        f.write(f"Edge-band attention ratio mean/std/median: {edge_mean:.6f} / {edge_std:.6f} / {edge_median:.6f}\n")
        f.write(f"Texture-attention corr mean/std/median: {tex_mean:.6f} / {tex_std:.6f} / {tex_median:.6f}\n")
        f.write(f"Representative sample ids: {selected_ids}\n")

    return {
        "sample_count": int(len(sample_df)),
        "spatial_layer_sizes": spatial_hw_shapes,
        "channel_layer_sizes": channel_shapes,
        "foreground_mean": fg_mean,
        "foreground_std": fg_std,
        "foreground_median": fg_median,
        "edge_mean": edge_mean,
        "edge_std": edge_std,
        "edge_median": edge_median,
        "texture_mean": tex_mean,
        "texture_std": tex_std,
        "texture_median": tex_median,
        "selected_ids": selected_ids
    }


train_generator = torch.Generator()
train_generator.manual_seed(SEED)
dataiter_tra = Data.DataLoader(dataset = train_set,
                           batch_size =TRAIN_BATCH_SIZE,
                           num_workers=0,
                           shuffle = True,
                           generator = train_generator)
dataiter_val = Data.DataLoader(dataset = val_set,
                           batch_size =TRAIN_BATCH_SIZE,
                           num_workers=0,
                           shuffle = False)

model_inverse = CNN(kernel_size=3,ratio=2)
model_forward=ForwardCNN(kernel_size=3,ratio=2)
# model_inverse=torch.load("0220-all/model/model_inverse-1.pkl")
# model_forward=torch.load("0220-all/model/model_forward-1.pkl")
model_inverse=model_inverse.to(device)
model_forward=model_forward.to(device)
# loss_fun = torch.nn.MSELoss()
# optimizer = torch.optim.Adam(model.parameters(), lr=0.0005)
loss_fun1 = torch.nn.MSELoss()
loss_fun2 = torch.nn.MSELoss()
# loss_fun1 = torch.nn.MSELoss()
# loss_fun2 = torch.nn.MSELoss()
# keep learning rate stable
opt1= torch.optim.Adam(model_inverse.parameters(),lr=0.001,weight_decay=1E-5)
opt2= torch.optim.Adam(model_forward.parameters(), lr=0.001,weight_decay=1E-5)

ckpt_inverse = MODEL_DIR / "model_inverse.pth"
ckpt_forward = MODEL_DIR / "model_forward.pth"
ckpt_train_state = MODEL_DIR / "train_state.pth"
if RUN_STAGE == "test":
    if (not ckpt_inverse.exists()) or (not ckpt_forward.exists()):
        raise FileNotFoundError(
            f"Missing trained checkpoints:\n{ckpt_inverse}\n{ckpt_forward}\nSet RUN_STAGE='train' to train first."
        )
    model_inverse, inv_load_info = load_inverse_model_with_cbam_guard(
        ckpt_inverse, device=device, kernel_size=3, ratio=2
    )
    state_forward = torch.load(str(ckpt_forward), map_location=device)
    if isinstance(state_forward, dict):
        model_forward.load_state_dict(state_forward)
    else:
        model_forward = state_forward
    model_forward = model_forward.to(device)
    print(
        "[InverseLoad] mode={}, spatial_attn={}, channel_attn={}".format(
            inv_load_info.get("mode"),
            inv_load_info.get("spatial"),
            inv_load_info.get("channel")
        )
    )
    if inv_load_info.get("mode") == "state_dict_into_current_cnn":
        print(
            "[InverseLoad] state_dict strict=False, missing_keys={}, unexpected_keys={}".format(
                inv_load_info.get("missing_keys"),
                inv_load_info.get("unexpected_keys")
            )
        )

# #changing learning rate with epoch
# lr=0.01
# opt1= torch.optim.Adam(model_inverse.parameters(),lr=lr,weight_decay=1E-5)
# opt2= torch.optim.Adam(model_forward.parameters(), lr=lr,weight_decay=1E-5)
# lr_list=[]
# lambda1=lambda epoch:0.95 ** epoch
# scheduler1=lr_scheduler.LambdaLR(opt1,lr_lambda=lambda1)
# scheduler2=lr_scheduler.LambdaLR(opt2,lr_lambda=lambda1)

# #zishiying
# lr=0.01
# opt1= torch.optim.Adam(model_inverse.parameters(),lr=lr)
# opt2= torch.optim.Adam(model_forward.parameters(), lr=lr)
# lr_list=[]
# scheduler1=lr_scheduler.ReduceLROnPlateau(opt1,mode='min',factor=0.5,patience=15000,verbose=False,threshold=1e-4,threshold_mode='rel',
#                                           cooldown=0,min_lr=0,eps=1e-8)
# scheduler2=lr_scheduler.ReduceLROnPlateau(opt2,mode='min',factor=0.5,patience=15000,verbose=False,threshold=1e-4,threshold_mode='rel',
#                                           cooldown=0,min_lr=0,eps=1e-8)
val_loss=[]
val_loss_total=[]
tra_loss=[]
tra_forward_loss=[]
r_2_total=[]
r_2_1=[]
r_2_train_total=[]
ssim_val1=[]
ssim_val_total=[]
best_val_r2 = -float("inf")
r2_drop_patience = 5
r2_drop_streak = 0
prev_val_r2 = None
start_epoch = 0

if RUN_STAGE == "train" and RESUME_TRAIN:
    if ckpt_train_state.exists():
        resume_state = torch.load(str(ckpt_train_state), map_location=device)
        if isinstance(resume_state, dict):
            if "model_inverse_state_dict" in resume_state:
                model_inverse.load_state_dict(resume_state["model_inverse_state_dict"])
            if "model_forward_state_dict" in resume_state:
                model_forward.load_state_dict(resume_state["model_forward_state_dict"])
            if "opt1_state_dict" in resume_state:
                opt1.load_state_dict(resume_state["opt1_state_dict"])
            if "opt2_state_dict" in resume_state:
                opt2.load_state_dict(resume_state["opt2_state_dict"])

            start_epoch = int(resume_state.get("epoch", -1)) + 1
            best_val_r2 = float(resume_state.get("best_val_r2", best_val_r2))
            r2_drop_streak = int(resume_state.get("r2_drop_streak", r2_drop_streak))
            prev_val_r2 = resume_state.get("prev_val_r2", prev_val_r2)

            val_loss = list(resume_state.get("val_loss", val_loss))
            val_loss_total = list(resume_state.get("val_loss_total", val_loss_total))
            tra_loss = list(resume_state.get("tra_loss", tra_loss))
            r_2_1 = list(resume_state.get("r_2_1", r_2_1))
            r_2_total = list(resume_state.get("r_2_total", r_2_total))
            r_2_train_total = list(resume_state.get("r_2_train_total", r_2_train_total))
            ssim_val1 = list(resume_state.get("ssim_val1", ssim_val1))
            ssim_val_total = list(resume_state.get("ssim_val_total", ssim_val_total))

            print(
                "Resume training from epoch {} (best_val_r2={:.6f}, drop_streak={})".format(
                    start_epoch, best_val_r2, r2_drop_streak
                )
            )
    else:
        print(f"Resume requested but checkpoint not found: {ckpt_train_state}. Start from scratch.")

epochs = 500 if RUN_STAGE == "train" else 0
for epoch in range(start_epoch, epochs):
    model_inverse.train()
    model_forward.train()

    train_loss_sum = 0.0
    train_count = 0
    train_true_list = []
    train_pred_list = []
    for step, (x, y) in enumerate(dataiter_tra):
        x = x.to(device)
        y = y.to(device)
        out = model_inverse(x)
        stru= model_forward(out)
        loss_inverse = loss_fun1(out, y)
        loss_forward = loss_fun2(stru,x)
        loss_all = loss_inverse+loss_forward
        opt1.zero_grad()
        opt2.zero_grad()
        # loss_inverse.backward(retain_graph=True)
        # loss_forward.backward(retain_graph=True)
        loss_all.backward()
        opt1.step()
        opt2.step()
        bs = x.size(0)
        train_loss_sum += float(loss_all.item()) * bs
        train_count += bs
        train_true_list.append(y.detach().cpu().numpy())
        train_pred_list.append(out.detach().cpu().numpy())

    if train_count == 0:
        raise RuntimeError("Empty training loader.")

    train_loss_epoch = train_loss_sum / train_count
    train_y_true = np.concatenate(train_true_list, axis=0)
    train_y_pred = np.concatenate(train_pred_list, axis=0)
    train_r2_epoch = r2_score(train_y_true, train_y_pred)
    print("epoch:{},loss:{:.4f}".format(epoch, train_loss_epoch))
    tra_loss.append(train_loss_epoch)
    r_2_train_total.append(train_r2_epoch)

    model_inverse.eval()
    model_forward.eval()
    val_loss_sum = 0.0
    val_count = 0
    val_ssim_sum = 0.0
    val_true_list = []
    val_pred_list = []
    with torch.no_grad():
        for step, (x1, y1) in enumerate(dataiter_val):
            x1 = x1.to(device)
            y1 = y1.to(device)
            validata_out = model_inverse(x1)
            val_stru= model_forward(validata_out)

            validata_inverse_loss = loss_fun1(validata_out, y1)
            validata_loss_forward=loss_fun2(val_stru,x1)
            validata_loss = float(validata_inverse_loss.item() + validata_loss_forward.item())
            ssim_val = float(ssim(x1, val_stru, size_average=True).item())

            bs = x1.size(0)
            val_loss_sum += validata_loss * bs
            val_ssim_sum += ssim_val * bs
            val_count += bs
            val_true_list.append(y1.detach().cpu().numpy())
            val_pred_list.append(validata_out.detach().cpu().numpy())

    if val_count == 0:
        raise RuntimeError("Empty validation loader.")

    val_loss1 = val_loss_sum / val_count
    val_ssim_epoch = val_ssim_sum / val_count
    val_y_true = np.concatenate(val_true_list, axis=0)
    val_y_pred = np.concatenate(val_pred_list, axis=0)
    r_2 = r2_score(val_y_true, val_y_pred)

    val_loss.append(val_loss1)
    val_loss_total.append(val_loss1)
    r_2_1.append(r_2)
    r_2_total.append(r_2)
    ssim_val1.append(val_ssim_epoch)
    ssim_val_total.append(val_ssim_epoch)
    if r_2 > best_val_r2:
        best_val_r2 = r_2
        torch.save(model_inverse, str(ckpt_inverse))
        torch.save(model_forward, str(ckpt_forward))

    if prev_val_r2 is not None and r_2 < prev_val_r2:
        r2_drop_streak += 1
    else:
        r2_drop_streak = 0
    prev_val_r2 = r_2

    if RUN_STAGE == "train":
        train_state = {
            "epoch": epoch,
            "model_inverse_state_dict": model_inverse.state_dict(),
            "model_forward_state_dict": model_forward.state_dict(),
            "opt1_state_dict": opt1.state_dict(),
            "opt2_state_dict": opt2.state_dict(),
            "best_val_r2": best_val_r2,
            "r2_drop_streak": r2_drop_streak,
            "prev_val_r2": prev_val_r2,
            "val_loss": val_loss,
            "val_loss_total": val_loss_total,
            "tra_loss": tra_loss,
            "r_2_1": r_2_1,
            "r_2_total": r_2_total,
            "r_2_train_total": r_2_train_total,
            "ssim_val1": ssim_val1,
            "ssim_val_total": ssim_val_total,
        }
        torch.save(train_state, str(ckpt_train_state))

    if r2_drop_streak >= r2_drop_patience:
        print(
            "Early stopping by R2 decrease: epoch={}, streak={}, current_r2={:.6f}, prev_r2={:.6f}".format(
                epoch, r2_drop_streak, r_2, r_2_total[-2]
            )
        )
        break

    if epoch % 20 == 0:
        print("epoch:{}, val_loss:{:.4f}, r2:{:.4f}, ssim:{:.4f}".format(epoch, val_loss1, r_2, val_ssim_epoch))
    # validata_data=validata_data.to(device)
    # validata_lable=validata_lable.to(device)
    # validata_out = model(validata_data)
    # validata_loss = loss_fun(validata_out, validata_lable)
    # validata_out=validata_out.cpu()
    # validata_lable=validata_lable.cpu()
    # validata_loss=validata_loss.cpu()
    # r_2 = r2_score(validata_lable.detach(), validata_out.detach())
    # val_loss.append(validata_loss.detach())
    # r_2_total.append(r_2)
# 浠呭湪璁粌妯″紡鍔犺浇鏈€浼樻ā鍨嬶紝鐒跺悗缁熶竴鎵ц娴嬭瘯
number = MODEL_TAG
if RUN_STAGE == "train":
    state_inverse = torch.load(str(ckpt_inverse), map_location=device)
    if isinstance(state_inverse, dict):
        model_inverse.load_state_dict(state_inverse)
    else:
        model_inverse = state_inverse
    model_inverse = model_inverse.to(device)
    state_forward = torch.load(str(ckpt_forward), map_location=device)
    if isinstance(state_forward, dict):
        model_forward.load_state_dict(state_forward)
    else:
        model_forward = state_forward
    model_forward = model_forward.to(device)

test_data, test_lable, test_out, test_stru, ssim_test = run_test_in_batches(
    model_inverse, model_forward, test_set, batch_size=TEST_BATCH_SIZE
)
print("Test SSIM:{:.4f}".format(ssim_test))

# Test-stage integration of attention interpretability analysis.
# This is append-only and does not alter original prediction tensors or metrics.
is_test_phase_now = True  # 杩欓噷宸茬粡澶勪簬缁熶竴娴嬭瘯闃舵
should_run_attn = ENABLE_ATTN_ANALYSIS and (is_test_phase_now or (not ATTN_ANALYSIS_ONLY_IN_TEST_MODE))

if should_run_attn:
    try:
        attn_summary = run_attention_interpretability_analysis(
            model_inverse, test_data, batch_size=TEST_BATCH_SIZE
        )
        print("[AttentionAnalysis] done. samples={}, fg_ratio_median={:.4f}, edge_band_ratio_median={:.4f}, texture_corr_mean={:.4f}, selected_ids={}".format(
            attn_summary["sample_count"],
            attn_summary["foreground_median"],
            attn_summary["edge_median"],
            attn_summary["texture_mean"],
            attn_summary["selected_ids"]
        ))
    except Exception as e:
        print(f"[AttentionAnalysis] skipped due to error: {e}")

toPIL=transforms.ToPILImage()
save_n = min(100, len(test_data))
for i in range(save_n):
    pic=toPIL(test_data[i])
    pic.save(str(TRUE_IMG_DIR / f"{i}.jpg"))
    pic1 = toPIL(test_stru[i])
    pic1.save(str(PRED_IMG_DIR / f"{i}.jpg"))
test_lable=test_lable.detach().numpy()
test_out=test_out.detach().numpy()

# factor1=10
# test_lable=[factor1 * i for i in list(test_lable.detach().numpy())]
# test_out=[factor1 * i for i in list(test_out.detach().numpy())]

# test_lable=mm.inverse_transform(test_lable.detach().numpy())
# test_out=mm.inverse_transform(test_out.detach().numpy())

# test_lable=ss.inverse_transform(test_lable.detach().numpy())
# test_out=ss.inverse_transform(test_out.detach().numpy())
pred_norm = np.clip(test_out, 0.0, 1.0)
true_norm = np.clip(test_lable, 0.0, 1.0)

pred_inc_angle = pred_norm[:,0] * 90.0
pred_azi_angle1 = pred_norm[:,1] * 360.0
pred_rad = pred_norm[:,2] * 200.0 + 50.0
pred_times = np.clip(np.round(pred_norm[:,3] * 3.0), 1, 3)
pred_azi_angle2 = pred_norm[:,4] * 360.0
pred_azi_angle3 = pred_norm[:,5] * 360.0

True_inc_angle = true_norm[:,0] * 90.0
True_azi_angle1 = true_norm[:,1] * 360.0
True_rad = true_norm[:,2] * 200.0 + 50.0
True_times = np.clip(np.round(true_norm[:,3] * 3.0), 1, 3)
True_azi_angle2 = true_norm[:,4] * 360.0
True_azi_angle3 = true_norm[:,5] * 360.0

val_loss1=np.ravel(val_loss)
r_2_total1=np.ravel(r_2_total)
train_loss=np.ravel(tra_loss)
train_r_2=np.ravel(r_2_train_total)

pred_inc_angle=np.around(pred_inc_angle)
pred_azi_angle1=np.around(pred_azi_angle1)
pred_azi_angle2=np.round(pred_azi_angle2)
pred_azi_angle3=np.round(pred_azi_angle3)

pred_rad=np.around(pred_rad)
pred_times=np.around(pred_times)
pred_azi_angle1[pred_azi_angle1<0]=0
pred_azi_angle2[pred_azi_angle2<0]=0
pred_azi_angle3[pred_azi_angle3<0]=0
pred_inc_angle[pred_inc_angle<0]=0

np.savetxt(str(RESULT_DIR / f"loss-test-{number}.txt"), val_loss1, fmt="%f")
np.savetxt(str(RESULT_DIR / f"r_2-test-{number}.txt"), r_2_total1, fmt="%f")
np.savetxt(str(RESULT_DIR / f"pred_inc_angle-{number}.txt"), pred_inc_angle, fmt="%d")
np.savetxt(str(RESULT_DIR / f"true_inc_angle-{number}.txt"), True_inc_angle, fmt="%d")
np.savetxt(str(RESULT_DIR / f"pred_azi_angle1-{number}.txt"), pred_azi_angle1, fmt="%d")
np.savetxt(str(RESULT_DIR / f"true_azi_angle1-{number}.txt"), True_azi_angle1, fmt="%d")
np.savetxt(str(RESULT_DIR / f"pred_azi_angle2-{number}.txt"), pred_azi_angle2, fmt="%d")
np.savetxt(str(RESULT_DIR / f"true_azi_angle2-{number}.txt"), True_azi_angle2, fmt="%d")
np.savetxt(str(RESULT_DIR / f"pred_azi_angle3-{number}.txt"), pred_azi_angle3, fmt="%d")
np.savetxt(str(RESULT_DIR / f"true_azi_angle3-{number}.txt"), True_azi_angle3, fmt="%d")
np.savetxt(str(RESULT_DIR / f"pred_rad-{number}.txt"), pred_rad, fmt="%d")
np.savetxt(str(RESULT_DIR / f"true_rad-{number}.txt"), True_rad, fmt="%d")
np.savetxt(str(RESULT_DIR / f"pred_times-{number}.txt"), pred_times, fmt="%d")
np.savetxt(str(RESULT_DIR / f"true_times-{number}.txt"), True_times, fmt="%d")

m=min(100, len(pred_inc_angle))
plt.title("Test")
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.xlabel("Sample")
plt.ylabel("Structure Parameters")
plt.subplot(3, 2, 1)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_inc_angle[0:m], label='Inc Pred')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_inc_angle[0:m], label='Inc True')
plt.legend()
plt.subplot(3, 2, 2)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_azi_angle1[0:m], label='Azi1 Pred')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_azi_angle1[0:m], label='Azi1 True')
plt.legend()
plt.subplot(3, 2, 3)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_rad[0:m], label='Rad Pred')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_rad[0:m], label='Rad True')
plt.legend()
plt.subplot(3, 2, 4)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_azi_angle2[0:m], label='Azi2 Pred')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_azi_angle2[0:m], label='Azi2 True')
plt.legend()
plt.subplot(3, 2, 5)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_azi_angle3[0:m], label='Azi3 Pred')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_azi_angle3[0:m], label='Azi3 True')
plt.legend()
plt.subplot(3, 2, 6)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_times[0:m], label='Times Pred')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_times[0:m], label='Times True')
plt.legend()
plt.show()

if openpyxl is not None:
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title='sheet1'
    col=('True-inc','Pred-inc','True-azi1','Pred-azi1','True-rad','Pred-rad','True-azi2','Pred-azi2','True-azi3','Pred-azi3','True-times','Pred-times')

    for index,item in enumerate(col):
        ws.cell(row=1,column=index+1,value=item)
        pass
    for i in range(m):
        ws.cell(row=i + 2, column=1, value=True_inc_angle[i])
        ws.cell(row=i + 2, column=2, value=pred_inc_angle[i])
        ws.cell(row=i + 2, column=3, value=True_azi_angle1[i])
        ws.cell(row=i + 2, column=4, value=pred_azi_angle1[i])
        ws.cell(row=i + 2, column=5, value=True_rad[i])
        ws.cell(row=i + 2, column=6, value=pred_rad[i])
        ws.cell(row=i + 2, column=7, value=True_azi_angle2[i])
        ws.cell(row=i + 2, column=8, value=pred_azi_angle2[i])
        ws.cell(row=i + 2, column=9, value=True_azi_angle3[i])
        ws.cell(row=i + 2, column=10, value=pred_azi_angle3[i])
        ws.cell(row=i + 2, column=11, value=pred_times[i])
        ws.cell(row=i + 2, column=12, value=True_times[i])
    wb.save(str(RESULT_DIR / f"test-pred-paramas-{number}.xlsx"))
else:
    print("openpyxl not installed, skip xlsx export.")








