import xlrd
import numpy as np
import torch.nn as nn
import torchvision.transforms as transforms
import torch
import openpyxl
import torch.utils.data as Data
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import StandardScaler,MinMaxScaler,Normalizer,scale
from sklearn.metrics import r2_score
from  matplotlib import pyplot as plt
import os
from PIL import Image
import matplotlib.image as mpimg
import torch.nn.functional as F
from math import exp
import numpy as np
import math
from torch.optim import lr_scheduler

#check if CUDA is available
use_cuda=torch.cuda.is_available()
print("cuda:",use_cuda)
device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')

path = "F:/zjl/matlab-program/20230210/all-sample-label.xlsx"
# path = "F:/zjl/attention-cnn/0209/test-pred-paramas2.xlsx"
data = xlrd.open_workbook(path)
table = data.sheet_by_index(0)
pictur_list = []
lable_list = []
lable_list_times=[]
loader=transforms.Compose([transforms.ToTensor()])
unloader=transforms.ToPILImage()
image_path="F:/zjl/matlab-program/20230210/total-sample-zip/"
# image_path="F:/zjl/attention-cnn/0209/true/"
for file in os.listdir(image_path):

    name,file_path=os.path.splitext(file)
    try:
        I = Image.open(image_path + name + ".jpg").convert('RGB')
        lable_list.append(table.row_values(int(name), start_colx=0, end_colx=6))

        # lable_list_times.append(table.row_values(int(name), start_colx=5, end_colx=6))
        # lable_list=loader(lable_list)
        I=loader(I)
        pictur_list.append(np.array(I))
    except FileNotFoundError:
        continue
# print("label_list:",lable_list)
picture_array = np.array(pictur_list)
# print("shape1",picture_array.shape)
lable_array = np.array(lable_list)
label1=lable_array[:,0]
label1=label1/90
# print("label1",label1)
label2=lable_array[:,1]
label2=label2/350
#
# print("label2",label2)
label3=lable_array[:,2]
label3=label3/270
label4=lable_array[:,3]
label4=label4/3
label5=lable_array[:,4]
label5=label5/350
label6=lable_array[:,4]
label6=label6/350

picture_array = picture_array.reshape(-1,3,138,80)
lable_array=np.array(list(zip(label1,label2,label3,label4,label5,label6)))
# print(lable_array)
# print("shape label",lable_array.shape)

# print("shape2",picture_array.shape)

index = [i for i in range(172447)]
np.random.shuffle(index)
lable_array = lable_array[index]
picture_array = picture_array[index]

picture_tensor = torch.from_numpy(picture_array).float()
# picture_tensor = picture_tensor.unsqueeze(1).float()
lable_tensor = torch.from_numpy(lable_array).float()


train_data = picture_tensor[0:120000]
train_lable = lable_tensor[0:120000]


validata_data = picture_tensor[120000:172347]
validata_lable = lable_tensor[120000:172347]


test_data = picture_tensor[172347:172447]
test_lable = lable_tensor[172347:172447]


class EarlyStopping(object):
    def __init__(self, monitor: str = 'val_loss', mode: str = 'min', patience: int = 1):
        """
        :param monitor: 要监测的指标，只有传入指标字典才会生效
        :param mode: 监测指标的模式，min 或 max
        :param patience: 最大容忍次数

        example:

        ```python
        # Initialize
        earlystopping = EarlyStopping(mode='max', patience=5)

        # call
        if earlystopping(val_accuracy):
           return;

        # save checkpoint

        state = {
            'model': model,
            'earlystopping': earlystopping.state_dict(),
            'optimizer': optimizer
        }

        torch.save(state, 'checkpoint.pth')

        checkpoint = torch.load('checkpoint.pth')
        earlystopping.load_state_dict(checkpoint['earlystopping'])
        ```
        """
        self.monitor = monitor
        self.mode = mode
        self.patience = patience
        self.__value = -math.inf if mode == 'max' else math.inf
        self.__times = 0

    def state_dict(self) -> dict:
        """:保存状态，以便下次加载恢复
        torch.save(state_dict, path)
        """
        return {
            'monitor': self.monitor,
            'mode': self.mode,
            'patience': self.patience,
            'value': self.__value,
            'times': self.__times
        }

    def load_state_dict(self, state_dict: dict):
        """:加载状态
        :param state_dict: 保存的状态
        """
        self.monitor = state_dict['monitor']
        self.mode = state_dict['mode']
        self.patience = state_dict['patience']
        self.__value = state_dict['value']
        self.__times = state_dict['times']

    def __call__(self, metrics) -> bool:
        """
        :param metrics: 指标字典或数值标量
        :return: 返回bool标量，True表示触发终止条件
        """
        if isinstance(metrics, dict):
            metrics = metrics[self.monitor]

        if (self.mode == 'min' and metrics <= self.__value) or (
                self.mode == 'max' and metrics >= self.__value):
            self.__value = metrics
            self.__times = 0
        else:
            self.__times += 1
        if self.__times >= self.patience:
            return True
        return False


# 计算一维的高斯分布向量
def gaussian(window_size, sigma):
    gauss = torch.Tensor([exp(-(x - window_size // 2) ** 2 / float(2 * sigma ** 2)) for x in range(window_size)])
    return gauss / gauss.sum()


# 创建高斯核，通过两个一维高斯分布向量进行矩阵乘法得到
# 可以设定channel参数拓展为3通道
def create_window(window_size, channel=1):
    _1D_window = gaussian(window_size, 1.5).unsqueeze(1)
    _2D_window = _1D_window.mm(_1D_window.t()).float().unsqueeze(0).unsqueeze(0)
    window = _2D_window.expand(channel, 1, window_size, window_size).contiguous()
    return window


# 计算SSIM
# 直接使用SSIM的公式，但是在计算均值时，不是直接求像素平均值，而是采用归一化的高斯核卷积来代替。
# 在计算方差和协方差时用到了公式Var(X)=E[X^2]-E[X]^2, cov(X,Y)=E[XY]-E[X]E[Y].
# 正如前面提到的，上面求期望的操作采用高斯核卷积代替。
def ssim(img1, img2, window_size=11, window=None, size_average=True, full=False, val_range=None):
    # Value range can be different from 255. Other common ranges are 1 (sigmoid) and 2 (tanh).
    if val_range is None:
        if torch.max(img1) > 128:
            max_val = 255
        else:
            max_val = 1

        if torch.min(img1) < -0.5:
            min_val = -1
        else:
            min_val = 0
        L = max_val - min_val
    else:
        L = val_range

    padd = 0
    (_, channel, height, width) = img1.size()
    if window is None:
        real_size = min(window_size, height, width)
        window = create_window(real_size, channel=channel).to(img1.device)

    mu1 = F.conv2d(img1, window, padding=padd, groups=channel)
    mu2 = F.conv2d(img2, window, padding=padd, groups=channel)

    mu1_sq = mu1.pow(2)
    mu2_sq = mu2.pow(2)
    mu1_mu2 = mu1 * mu2

    sigma1_sq = F.conv2d(img1 * img1, window, padding=padd, groups=channel) - mu1_sq
    sigma2_sq = F.conv2d(img2 * img2, window, padding=padd, groups=channel) - mu2_sq
    sigma12 = F.conv2d(img1 * img2, window, padding=padd, groups=channel) - mu1_mu2

    C1 = (0.01 * L) ** 2
    C2 = (0.03 * L) ** 2

    v1 = 2.0 * sigma12 + C2
    v2 = sigma1_sq + sigma2_sq + C2
    cs = torch.mean(v1 / v2)  # contrast sensitivity

    ssim_map = ((2 * mu1_mu2 + C1) * v1) / ((mu1_sq + mu2_sq + C1) * v2)

    if size_average:
        ret = ssim_map.mean()
    else:
        ret = ssim_map.mean(1).mean(1).mean(1)

    if full:
        return ret, cs
    return ret

class ChannelAttention(nn.Module):
    def __init__(self, in_channel, ratio=16):
        """ 通道注意力机制 同最大池化和平均池化两路分别提取信息，后共用一个多层感知机mlp,再将二者结合

        :param in_channel: 输入通道
        :param ratio: 通道降低倍率
        """
        super(ChannelAttention, self).__init__()
        # 平均池化
        self.avg_pool = nn.AdaptiveAvgPool2d(1)
        # 最大池化
        self.max_pool = nn.AdaptiveMaxPool2d(1)

        # 通道先降维后恢复到原来的维数
        self.fc1 = nn.Conv2d(in_channel, in_channel // ratio, 1, bias=False)
        self.relu1 = nn.ReLU()
        self.fc2 = nn.Conv2d(in_channel // ratio, in_channel, 1, bias=False)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        # print("x.shape",x.shape)
        # 平均池化
        # avg_out = self.fc2(self.relu1(self.fc1(self.avg_pool(x))))
        # 最大池化
        # max_out = self.fc2(self.relu1(self.fc1(self.max_pool(x))))
        # out = avg_out + max_out
        # return x*self.sigmoid(out)

        # 平均池化一支 (2,512,8,8) -> (2,512,1,1) -> (2,512/ration,1,1) -> (2,512,1,1)
        # (2,512,8,8) -> (2,512,1,1)
        avg = self.avg_pool(x)
        # 多层感知机mlp (2,512,8,8) -> (2,512,1,1) -> (2,512/ration,1,1) -> (2,512,1,1)
        # (2,512,1,1) -> (2,512/ratio,1,1)
        avg = self.fc1(avg)
        avg = self.relu1(avg)
        # (2,512/ratio,1,1) -> (2,512,1,1)
        avg_out = self.fc2(avg)

        # 最大池化一支
        # (2,512,8,8) -> (2,512,1,1)
        max = self.max_pool(x)
        # 多层感知机
        # (2,512,1,1) -> (2,512/ratio,1,1)
        max = self.fc1(max)
        max = self.relu1(max)
        # (2,512/ratio,1,1) -> (2,512,1,1)
        max_out = self.fc2(max)

        # (2,512,1,1) + (2,512,1,1) -> (2,512,1,1)
        out = avg_out + max_out
        output=self.sigmoid(out)
        # print("output.shape",output.shape)
        return x*output

class SpatialAttention(nn.Module):
    def __init__(self, kernel_size=7):
        """ 空间注意力机制 将通道维度通过最大池化和平均池化进行压缩，然后合并，再经过卷积和激活函数，结果和输入特征图点乘

        :param kernel_size: 卷积核大小
        """
        super(SpatialAttention, self).__init__()

        assert kernel_size in (3, 7), 'kernel size must be 3 or 7'
        padding = 3 if kernel_size == 7 else 1

        self.conv1 = nn.Conv2d(2, 1, kernel_size, padding=padding, bias=False)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        # print('x shape', x.shape)
        # (2,512,8,8) -> (2,1,8,8)
        avg_out = torch.mean(x, dim=1, keepdim=True)
        # (2,512,8,8) -> (2,1,8,8)
        max_out, _ = torch.max(x, dim=1, keepdim=True)
        # (2,1,8,8) + (2,1,8,8) -> (2,2,8,8)
        cat = torch.cat([avg_out, max_out], dim=1)
        # (2,2,8,8) -> (2,1,8,8)
        out = self.conv1(cat)
        output=self.sigmoid(out)
        # print("output.shape", output.shape)
        return x*output

# Classes to re-use window
class SSIM(torch.nn.Module):
    def __init__(self, window_size=11, size_average=True, val_range=None):
        super(SSIM, self).__init__()
        self.window_size = window_size
        self.size_average = size_average
        self.val_range = val_range

        # Assume 1 channel for SSIM
        self.channel = 1
        self.window = create_window(window_size)

    def forward(self, img1, img2):
        (_, channel, _, _) = img1.size()

        if channel == self.channel and self.window.dtype == img1.dtype:
            window = self.window
        else:
            window = create_window(self.window_size, channel).to(img1.device).type(img1.dtype)
            self.window = window
            self.channel = channel

        return ssim(img1, img2, window=window, window_size=self.window_size, size_average=self.size_average)
#搭建网络
class CNN(nn.Module):
    def __init__(self,ratio,kernel_size):
        super(CNN,self).__init__()
        # 组合的卷积模块
        self.conv = nn.Sequential(
            nn.Conv2d(3, 10, 3,padding=1), #卷积层，输入通道数为3，输出通道数为10，卷积核大小为5X5
            nn.BatchNorm2d(10),
            nn.ReLU(), #激活函数层
            nn.MaxPool2d(2, 2), #最大池化层，卷积核大小为2X2，图片减小一半
            # SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(10, 16, 3,padding=1), #卷积层，输入通道数为10，输出通道数为16，卷积核大小为5X5
            nn.BatchNorm2d(16),
            nn.ReLU(),
            # nn.MaxPool2d(2, 2),
            # SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(16, 21, 3,padding=1),
            nn.BatchNorm2d(21),
            nn.ReLU(),
            # nn.MaxPool2d(2, 2),
            # SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(21, 24, 3,padding=1),
            nn.BatchNorm2d(24),
            nn.ReLU(),

            # SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(24, 32, 3,padding=1),
            nn.BatchNorm2d(32),
            nn.ReLU(),
            # SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(32, 36, 3,padding=1),
            nn.BatchNorm2d(36),
            nn.ReLU(),
            # nn.MaxPool2d(2, 2)
            # SpatialAttention(kernel_size=kernel_size),
        )
        self.fc = nn.Sequential(
            nn.Linear(36*69*40, 120), #线性层
            torch.nn.BatchNorm1d(120), #归一化层
            nn.Dropout(0.2),
            nn.ReLU(),#激活函数层
            nn.Linear(120, 84),
            torch.nn.BatchNorm1d(84),
            nn.Dropout(0.2),
            nn.ReLU(),
            nn.Linear(84, 20),
            torch.nn.BatchNorm1d(20),
            nn.Dropout(0.2),
            nn.ReLU(),
            nn.Linear(20, 6), #有三种标签，分别预测

        )

    def forward(self, x):
        y = self.conv(x)
        output = self.fc(y.view(x.shape[0], -1))
        return output
class ForwardCNN(nn.Module):
    def __init__(self,kernel_size,ratio):
        super( ForwardCNN,self).__init__()
        self.fc = nn.Sequential(
            nn.Linear(6, 20),  # 有三种标签，分别预测
            nn.Linear(20, 84),
            torch.nn.BatchNorm1d(84),
            # nn.Dropout(0.2),
            nn.ReLU(),
            nn.Linear(84, 120),
            torch.nn.BatchNorm1d(120),
            # nn.Dropout(0.2),
            nn.ReLU(),
            nn.Linear(120,36 * 69 * 40),  # 线性层
            torch.nn.BatchNorm1d(36 * 69 * 40),  # 归一化层
            nn.Dropout(0.2),
            nn.ReLU(),  # 激活函数层
        )
        # 组合的卷积模块
        self.conv = nn.Sequential(
            nn.Upsample(scale_factor=2),
            nn.Conv2d(36, 32, 3,padding=1),  # 卷积层，输入通道数为3，输出通道数为10，卷积核大小为5X5
            nn.BatchNorm2d(32),
            nn.ReLU(),  # 激活函数层
            # SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(32, 24, 3,padding=1),  # 卷积层，输入通道数为3，输出通道数为10，卷积核大小为5X5
            nn.BatchNorm2d(24),
            nn.ReLU(),  # 激活函数层
            # SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(24, 21, 3,padding=1), #卷积层，输入通道数为3，输出通道数为10，卷积核大小为5X5
            nn.BatchNorm2d(21),
            nn.ReLU(), #激活函数层
            # nn.MaxUnpool2d(2, 2), #最大池化层，卷积核大小为2X2，图片减小一半
            # SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(21, 16, 3,padding=1), #卷积层，输入通道数为10，输出通道数为16，卷积核大小为5X5
            nn.BatchNorm2d(16),
            nn.ReLU(),
            # nn.MaxUnpool2d(2, 2),
            # SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(16, 10, 3,padding=1),
            nn.BatchNorm2d(10),
            nn.ReLU(),
            # nn.MaxUnpool2d(2, 2),
            # SpatialAttention(kernel_size=kernel_size),

            nn.Conv2d(10, 3, 3,padding=1),
            nn.BatchNorm2d(3),
            nn.ReLU(),
            # nn.LogSoftmax()
            # nn.MaxUnpool2d(2, 2)
            # SpatialAttention(kernel_size=kernel_size),
        )

    def forward(self, x):
        y = self.fc(x.view(x.shape[0], -1))
        y=torch.unsqueeze(y,dim=2)
        y=torch.unsqueeze(y,dim=3)
        y=torch.reshape(y,(-1,36,69,40))
        output= self.conv(y)
        return output

train_set = Data.TensorDataset(train_data,train_lable)
val_set=Data.TensorDataset(validata_data,validata_lable)
dataiter_tra = Data.DataLoader(dataset = train_set,
                           batch_size =300,
                           num_workers=0,
                           shuffle = True)
dataiter_val = Data.DataLoader(dataset = val_set,
                           batch_size =300,
                           num_workers=0,
                           shuffle = True)

model_inverse = CNN(kernel_size=3,ratio=2)
model_forward=ForwardCNN(kernel_size=3,ratio=2)
# model_inverse=torch.load("0220-all/model/model_inverse-1.pkl")
# model_forward=torch.load("0220-all/model/model_forward-1.pkl")
model_inverse=model_inverse.to(device)
model_forward=model_forward.to(device)
# loss_fun = torch.nn.MSELoss()
# optimizer = torch.optim.Adam(model.parameters(), lr=0.0005)
loss_fun1 = torch.nn.MSELoss()
loss_fun2 = torch.nn.MSELoss()
# loss_fun1 = torch.nn.MSELoss()
# loss_fun2 = torch.nn.MSELoss()
# keep learning rate stable
opt1= torch.optim.Adam(model_inverse.parameters(),lr=0.001,weight_decay=1E-5)
opt2= torch.optim.Adam(model_forward.parameters(), lr=0.001,weight_decay=1E-5)

# #changing learning rate with epoch
# lr=0.01
# opt1= torch.optim.Adam(model_inverse.parameters(),lr=lr,weight_decay=1E-5)
# opt2= torch.optim.Adam(model_forward.parameters(), lr=lr,weight_decay=1E-5)
# lr_list=[]
# lambda1=lambda epoch:0.95 ** epoch
# scheduler1=lr_scheduler.LambdaLR(opt1,lr_lambda=lambda1)
# scheduler2=lr_scheduler.LambdaLR(opt2,lr_lambda=lambda1)

# #zishiying
# lr=0.01
# opt1= torch.optim.Adam(model_inverse.parameters(),lr=lr)
# opt2= torch.optim.Adam(model_forward.parameters(), lr=lr)
# lr_list=[]
# scheduler1=lr_scheduler.ReduceLROnPlateau(opt1,mode='min',factor=0.5,patience=15000,verbose=False,threshold=1e-4,threshold_mode='rel',
#                                           cooldown=0,min_lr=0,eps=1e-8)
# scheduler2=lr_scheduler.ReduceLROnPlateau(opt2,mode='min',factor=0.5,patience=15000,verbose=False,threshold=1e-4,threshold_mode='rel',
#                                           cooldown=0,min_lr=0,eps=1e-8)
val_loss=[]
val_loss_total=[]
tra_loss=[]
tra_forward_loss=[]
r_2_total=[]
r_2_1=[]
r_2_train=[]
r_2_train_total=[]
ssim_val1=[]
ssim_val_total=[]
epochs =1000
# 初始化，监测模式为最大，最多容忍5次
early_stop = EarlyStopping(mode='max', patience=3)
for epoch in range(epochs):

    for step, (x, y) in enumerate(dataiter_tra):
        x = x.to(device)
        y = y.to(device)
        out = model_inverse(x)
        stru= model_forward(out)
        loss_inverse = loss_fun1(out, y)
        loss_forward = loss_fun2(stru,x)
        loss_all = loss_inverse+loss_forward
        opt1.zero_grad()
        opt2.zero_grad()
        # loss_inverse.backward(retain_graph=True)
        # loss_forward.backward(retain_graph=True)
        loss_all.backward()
        opt1.step()
        opt2.step()
        y = y.cpu()
        x = x.cpu()
        out = out.cpu()
        r_2_t = r2_score(y.detach(), out.detach())
        r_2_train.append(r_2_t)
        # scheduler1.step(loss_forward)
        # scheduler2.step(loss_all)
        # print("loss",loss_all.size())
    print("epoch:{},loss:{:.4f}".format(epoch, np.mean(loss_all.item())))
    #     loss_all=loss_all.cpu()
    #     tra_loss.append(loss_all.detach())
    #     # print("epoch:{},step:{},loss:{}".format(epoch, step, loss))
    tra_loss.append(np.mean(loss_all.item()))
    r_2_train_total.append(np.mean(r_2_train))
    model_inverse.eval()
    model_forward.eval()
    for step, (x1, y1) in enumerate(dataiter_val):
        # move to GPU
        x1 = x1.to(device)
        y1 = y1.to(device)
        validata_out = model_inverse(x1)
        val_stru= model_forward(validata_out)

        validata_inverse_loss = loss_fun1(validata_out, y1)
        validata_loss_forward=loss_fun2(val_stru,x1)
        validata_loss = validata_inverse_loss.item() + validata_loss_forward.item()
        val_loss.append(validata_loss)
        # if epoch%50==0:
        y1 = y1.cpu()
        x1=x1.cpu()
        val_stru=val_stru.cpu()
        validata_out = validata_out.cpu()
        r_2_ = r2_score(y1.detach(), validata_out.detach())

        r_2_1.append(r_2_)
        ssim_val = ssim(x1.detach(), val_stru.detach(), size_average=True)
        ssim_val1.append(ssim_val)

    val_loss1 = np.mean(val_loss)
    ssim_val_total.append(np.mean(ssim_val1))
    val_loss_total.append(val_loss1)
    r_2 = np.mean(r_2_1)
    r_2_total.append(r_2)
    if epoch>200:
        # 如果触发终止条件，就结束训练
        if early_stop(r_2):
            print("Early stopping")
            break
    if epoch%50==0:
        print("epoch:{},验证损失:{:.4f},绝对系数:{:.4f},SSIM距离:{:.4f}".format(epoch, np.mean(val_loss), r_2, np.mean(ssim_val1)))
    # validata_data=validata_data.to(device)
    # validata_lable=validata_lable.to(device)
    # validata_out = model(validata_data)
    # validata_loss = loss_fun(validata_out, validata_lable)
    # validata_out=validata_out.cpu()
    # validata_lable=validata_lable.cpu()
    # validata_loss=validata_loss.cpu()
    # r_2 = r2_score(validata_lable.detach(), validata_out.detach())
    # val_loss.append(validata_loss.detach())
    # r_2_total.append(r_2)
    # print("epoch:{},验证损失:{},绝对系数:{}".format(epoch,validata_loss, r_2))
number="cnn-"+str(1)
torch.save(model_inverse,"20240112/model/model_inverse-"+number+"pkl")
torch.save(model_forward,"20240112/model/model_forward-"+number+"pkl")

test_lable=test_lable.to(device)
test_data=test_data.to(device)
test_out = model_inverse(test_data)
test_stru = model_forward(test_out)
test_out=test_out.cpu()
test_lable=test_lable.cpu()
test_stru=test_stru.cpu()
test_data=test_data.cpu()
ssim_test = ssim(test_stru.detach(), test_data.detach(),size_average=True)
print("测试的SSIM距离:{:.4f}".format(ssim_test))

true_pic="20240112/true/"
pre_pic="20240112/pre/"
toPIL=transforms.ToPILImage()
for i in range (100):
    pic=toPIL(test_data[i])
    pic.save(true_pic+str(i) +'.jpg')
    pic1 = toPIL(test_stru[i])
    pic1.save(pre_pic + str(i) + '.jpg')
test_lable=test_lable.detach().numpy()
test_out=test_out.detach().numpy()

# factor1=10
# test_lable=[factor1 * i for i in list(test_lable.detach().numpy())]
# test_out=[factor1 * i for i in list(test_out.detach().numpy())]

# test_lable=mm.inverse_transform(test_lable.detach().numpy())
# test_out=mm.inverse_transform(test_out.detach().numpy())

# test_lable=ss.inverse_transform(test_lable.detach().numpy())
# test_out=ss.inverse_transform(test_out.detach().numpy())
pred_inc_angle=test_out[:,0]*90
pred_azi_angle1=test_out[:,1]*350
pred_rad=test_out[:,2]*270
pred_times=test_out[:,3]*3
pred_azi_angle2=test_out[:,4]*350
pred_azi_angle3=test_out[:,5]*350

True_inc_angle=test_lable[:,0]*90
True_azi_angle1=test_lable[:,1]*350
True_rad=test_lable[:,2]*270
True_times=test_lable[:,3]*3
True_azi_angle2=test_lable[:,4]*350
True_azi_angle3=test_lable[:,5]*350

val_loss1=np.ravel(val_loss)
r_2_total1=np.ravel(r_2_total)
train_loss=np.ravel(tra_loss)
train_r_2=np.ravel(r_2_train_total)

pred_inc_angle=np.around(pred_inc_angle)
pred_azi_angle1=np.around(pred_azi_angle1)
pred_azi_angle2=np.round(pred_azi_angle2)
pred_azi_angle3=np.round(pred_azi_angle3)

pred_rad=np.around(pred_rad)
pred_times=np.around(pred_times)
pred_azi_angle1[pred_azi_angle1<0]=0
pred_azi_angle2[pred_azi_angle2<0]=0
pred_azi_angle3[pred_azi_angle3<0]=0
pred_inc_angle[pred_inc_angle<0]=0

#保存参数
# np.savetxt("F:/zjl/attention-cnn/0116/result/ssim-val.txt", ssim_val_total, fmt="%f")
# np.savetxt("F:/zjl/attention-cnn/0116/result/loss-train.txt", train_loss, fmt="%f")
# np.savetxt("F:/zjl/attention-cnn/0116/result/r_2-train.txt", train_r_2, fmt="%f")
np.savetxt("F:/zjl/attention-cnn/20240112/result/loss-test-"+number+".txt", val_loss1, fmt="%f")
np.savetxt("F:/zjl/attention-cnn/20240112/result/r_2-test-"+number+".txt", r_2_total1, fmt="%f")
np.savetxt("F:/zjl/attention-cnn/20240112/result/pred_inc_angle-"+number+".txt", pred_inc_angle, fmt="%d")
np.savetxt("F:/zjl/attention-cnn/20240112/result/true_inc_angle-"+number+".txt", True_inc_angle, fmt="%d")
np.savetxt("F:/zjl/attention-cnn/20240112/result/pred_azi_angle1-"+number+".txt", pred_azi_angle1, fmt="%d")
np.savetxt("F:/zjl/attention-cnn/20240112/result/true_azi_angle1-"+number+".txt", True_azi_angle1, fmt="%d")
np.savetxt("F:/zjl/attention-cnn/20240112/result/pred_azi_angle2-"+number+".txt", pred_azi_angle2, fmt="%d")
np.savetxt("F:/zjl/attention-cnn/20240112/result/true_azi_angle2-"+number+".txt", True_azi_angle2, fmt="%d")
np.savetxt("F:/zjl/attention-cnn/20240112/result/pred_azi_angle3-"+number+".txt", pred_azi_angle3, fmt="%d")
np.savetxt("F:/zjl/attention-cnn/20240112/result/true_azi_angle3-"+number+".txt", True_azi_angle3, fmt="%d")
np.savetxt("F:/zjl/attention-cnn/20240112/result/pred_rad-"+number+".txt", pred_rad, fmt="%d")
np.savetxt("F:/zjl/attention-cnn/20240112/result/true_rad-"+number+".txt", True_rad, fmt="%d")
np.savetxt("F:/zjl/attention-cnn/20240112/result/pred_times-"+number+".txt", pred_times, fmt="%d")
np.savetxt("F:/zjl/attention-cnn/20240112/result/true_times-"+number+".txt", True_times, fmt="%d")

#画图
m=100
plt.title("测试")
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.xlabel("样本数")
plt.ylabel("结构参数")
plt.subplot(3, 2, 1)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1),pred_inc_angle[0:m], label='入射角预测')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_inc_angle[0:m], label='入射角真实')
plt.legend()
plt.subplot(3, 2, 2)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_azi_angle1[0:m], label='方位角1预测')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_azi_angle1[0:m], label='方位角1真实')
plt.legend()
plt.subplot(3, 2, 3)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_rad[0:m], label='半径预测')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_rad[0:m], label='半径真实')
plt.legend()
plt.subplot(3, 2, 4)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_azi_angle2[0:m], label='方位角2预测')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_azi_angle2[0:m], label='方位角2真实')
plt.legend()
plt.subplot(3, 2, 5)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_azi_angle3[0:m], label='方位角3预测')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_azi_angle3[0:m], label='方位角3真实')
plt.legend()
plt.subplot(3, 2, 6)
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), pred_times[0:m], label='Times 预测')
plt.plot(np.array([i for i in range(1, m+1)]).reshape(m, 1), True_times[0:m], label='Times 真实')
plt.legend()


plt.show()

wb=openpyxl.Workbook()
ws=wb.active
ws.title='sheet1'
col=('True-inc','Pred-inc','True-azi1','Pred-azi1','True-rad','Pred-rad','True-azi2','Pred-azi2','True-azi3','Pred-azi3','True-times','Pred-times')

for index,item in enumerate(col):
    ws.cell(row=1,column=index+1,value=item)
    pass
for i in range(100):
    ws.cell(row=i + 2, column=1, value=True_inc_angle[i])
    ws.cell(row=i + 2, column=2, value=pred_inc_angle[i])
    ws.cell(row=i + 2, column=3, value=True_azi_angle1[i])
    ws.cell(row=i + 2, column=4, value=pred_azi_angle1[i])
    ws.cell(row=i + 2, column=5, value=True_rad[i])
    ws.cell(row=i + 2, column=6, value=pred_rad[i])
    ws.cell(row=i + 2, column=7, value=True_azi_angle2[i])
    ws.cell(row=i + 2, column=8, value=pred_azi_angle2[i])
    ws.cell(row=i + 2, column=9, value=True_azi_angle3[i])
    ws.cell(row=i + 2, column=10, value=pred_azi_angle3[i])
    ws.cell(row=i + 2, column=11, value=pred_times[i])
    ws.cell(row=i + 2, column=12, value=True_times[i])
wb.save("F:/zjl/attention-cnn/20240112/result/test-pred-paramas-"+number+".xlsx")








