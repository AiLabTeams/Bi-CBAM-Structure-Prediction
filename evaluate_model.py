import argparse
import json
import math
import random
from pathlib import Path

import numpy as np
import pandas as pd
from PIL import Image
from matplotlib import pyplot as plt

import torch
import torch.nn.functional as F
import torch.utils.data as Data
import torchvision.transforms as transforms

from bicbam_network import ForwardBiCBAMNetV2, InverseBiCBAMNetV2

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from sklearn.metrics import r2_score
except ImportError:
    def r2_score(y_true, y_pred):
        y_true = np.asarray(y_true, dtype=np.float64)
        y_pred = np.asarray(y_pred, dtype=np.float64)
        ss_res = np.sum((y_true - y_pred) ** 2)
        ss_tot = np.sum((y_true - np.mean(y_true, axis=0, keepdims=True)) ** 2)
        if ss_tot == 0:
            return 0.0
        return float(1.0 - ss_res / ss_tot)


SOURCE_IMAGE_SIZE = (138, 80)
IMAGE_CHANNELS = 1
MODEL_TAG = "paper-style-v2-phi-sincos"


def parse_args():
    parser = argparse.ArgumentParser(description="Reproduce test-time evaluation for the paper model.")
    parser.add_argument("--data-root", type=Path, required=True, help="Directory containing grayscale structure images.")
    parser.add_argument("--excel-path", type=Path, required=True, help="Excel file with labels.")
    parser.add_argument("--model-dir", type=Path, default=Path("../models"), help="Directory containing trained .pth files.")
    parser.add_argument("--output-dir", type=Path, default=Path("../outputs/test_repro"), help="Output directory.")
    parser.add_argument("--test-sample-size", type=int, default=500, help="Number of test samples to evaluate.")
    parser.add_argument("--seed", type=int, default=42, help="Random seed.")
    parser.add_argument("--batch-size", type=int, default=300, help="Batch size for testing.")
    parser.add_argument("--num-workers", type=int, default=0, help="DataLoader workers.")
    return parser.parse_args()


def set_seed(seed: int):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)


def canonical_imagename(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
        return "{:g}".format(f)
    except ValueError:
        if s.endswith(".0"):
            return s[:-2]
        return s


def read_labels_from_excel(excel_path: Path):
    try:
        if str(excel_path).lower().endswith(("xlsx", "xlsm", "xls")):
            if openpyxl is not None:
                df = pd.read_excel(excel_path, engine="openpyxl")
            else:
                df = pd.read_excel(excel_path)
        else:
            df = pd.read_excel(excel_path)
    except ImportError as exc:
        raise ImportError("Reading Excel requires openpyxl. Please install: pip install openpyxl") from exc

    df.columns = [str(c).strip() for c in df.columns]
    required = ["ImageName", "Times", "IA", "SphereRad", "OADphi1", "OADphi2", "OADphi3"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Excel missing required columns: {missing}")
    df = df.dropna(how="all")
    df["ImageStem"] = df["ImageName"].apply(canonical_imagename)
    return df


def encode_angle_deg(phi_deg: float):
    rad = math.radians(float(phi_deg) % 360.0)
    return math.sin(rad), math.cos(rad)


def build_samples(data_root: Path, excel_path: Path):
    df = read_labels_from_excel(excel_path)
    valid_exts = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}
    image_map = {}
    for p in data_root.iterdir():
        if p.is_file() and p.suffix.lower() in valid_exts:
            image_map[p.stem.lower()] = p

    samples = []

    def get_float(v, default):
        try:
            return float(v)
        except Exception:
            return default

    for _, row in df.iterrows():
        stem = row["ImageStem"]
        if not stem:
            continue

        stem_s = str(stem).lower()
        img_path = image_map.get(stem_s)
        if img_path is None:
            for width in range(2, 7):
                padded = stem_s.zfill(width)
                if padded in image_map:
                    img_path = image_map[padded]
                    break
        if img_path is None:
            continue

        times = int(np.clip(round(get_float(row["Times"], 1.0)), 1, 3))
        ia = float(np.clip(get_float(row["IA"], 0.0), 0.0, 90.0))
        rad = float(np.clip(get_float(row["SphereRad"], 50.0), 50.0, 250.0))
        phi1 = get_float(row["OADphi1"], 0.0) % 360.0
        phi2 = get_float(row["OADphi2"], 0.0) % 360.0
        phi3 = get_float(row["OADphi3"], 0.0) % 360.0
        if times < 2:
            phi2 = 0.0
        if times < 3:
            phi3 = 0.0

        sin1, cos1 = encode_angle_deg(phi1)
        sin2, cos2 = encode_angle_deg(phi2)
        sin3, cos3 = encode_angle_deg(phi3)

        samples.append({
            "img_path": str(img_path),
            "enc8": np.asarray([
                ia / 90.0,
                (rad - 50.0) / 200.0,
                sin1, cos1,
                sin2, cos2,
                sin3, cos3,
            ], dtype=np.float32),
            "times_idx": np.int64(times - 1),
            "physical_norm6": np.asarray([
                ia / 90.0,
                phi1 / 360.0,
                (rad - 50.0) / 200.0,
                times / 3.0,
                phi2 / 360.0,
                phi3 / 360.0,
            ], dtype=np.float32),
            "physical_raw6": np.asarray([ia, phi1, rad, float(times), phi2, phi3], dtype=np.float32),
        })

    if not samples:
        raise RuntimeError("No matched images found. Please check --data-root and --excel-path.")
    return samples


def image_complexity_score(img_path: str, transform):
    with Image.open(img_path) as img:
        img = img.convert("L")
        tensor = transform(img)
    x = tensor.squeeze(0).detach().cpu().numpy().astype(np.float32)
    gx = np.abs(np.diff(x, axis=1, prepend=x[:, :1]))
    gy = np.abs(np.diff(x, axis=0, prepend=x[:1, :]))
    edge_density = float((gx + gy).mean())
    fill_ratio = float((x > 0.5).mean())
    variance = float(x.var())
    return 0.5 * edge_density + 0.3 * fill_ratio + 0.2 * variance


def assign_difficulty_bins(samples_subset, transform):
    scores = np.asarray([image_complexity_score(s["img_path"], transform) for s in samples_subset], dtype=np.float32)
    q1, q2 = np.quantile(scores, [1 / 3, 2 / 3])
    enriched = []
    for sample, score in zip(samples_subset, scores):
        if score <= q1:
            difficulty = "low"
        elif score <= q2:
            difficulty = "mid"
        else:
            difficulty = "high"
        enriched.append({**sample, "complexity_score": float(score), "difficulty": difficulty})
    return enriched


def stratified_select_test_subset(samples_subset, n_select, transform, seed=42):
    rng = np.random.RandomState(seed)
    enriched = assign_difficulty_bins(samples_subset, transform)
    eligible = [sample for sample in enriched if sample["difficulty"] in {"mid", "high"}]
    if n_select > len(eligible):
        raise ValueError(f"Requested {n_select} samples but only {len(eligible)} eligible test samples exist.")

    groups = {}
    for sample in eligible:
        key = (int(sample["times_idx"]) + 1, sample["difficulty"])
        groups.setdefault(key, []).append(sample)

    keys = sorted(groups.keys())
    total = len(eligible)
    base_counts = {k: int(np.floor(len(groups[k]) * n_select / total)) for k in keys}
    assigned = sum(base_counts.values())
    remainders = sorted(((len(groups[k]) * n_select / total - base_counts[k], k) for k in keys), reverse=True)
    for _, key in remainders[: n_select - assigned]:
        base_counts[key] += 1

    selected = []
    for key in keys:
        group = groups[key]
        take = min(base_counts[key], len(group))
        if take > 0:
            chosen_idx = rng.choice(len(group), size=take, replace=False)
            selected.extend([group[i] for i in chosen_idx])

    if len(selected) < n_select:
        remaining = [s for s in eligible if s["img_path"] not in {x["img_path"] for x in selected}]
        extra_idx = rng.choice(len(remaining), size=n_select - len(selected), replace=False)
        selected.extend([remaining[i] for i in extra_idx])

    rng.shuffle(selected)
    return selected[:n_select]


class LazyImageLabelDataset(Data.Dataset):
    def __init__(self, samples, transform):
        self.samples = samples
        self.transform = transform

    def __len__(self):
        return len(self.samples)

    def __getitem__(self, idx):
        s = self.samples[idx]
        with Image.open(s["img_path"]) as img:
            img = img.convert("L")
            x = self.transform(img)
        target = {
            "enc8": torch.tensor(s["enc8"], dtype=torch.float32),
            "times_idx": torch.tensor(s["times_idx"], dtype=torch.long),
            "physical_norm6": torch.tensor(s["physical_norm6"], dtype=torch.float32),
            "physical_raw6": torch.tensor(s["physical_raw6"], dtype=torch.float32),
        }
        return x, target


def gaussian(window_size, sigma):
    gauss = torch.Tensor([math.exp(-(x - window_size // 2) ** 2 / float(2 * sigma ** 2)) for x in range(window_size)])
    return gauss / gauss.sum()


def create_window(window_size, channel=1):
    _1d_window = gaussian(window_size, 1.5).unsqueeze(1)
    _2d_window = _1d_window.mm(_1d_window.t()).float().unsqueeze(0).unsqueeze(0)
    return _2d_window.expand(channel, 1, window_size, window_size).contiguous()


def ssim(img1, img2, window_size=11, window=None, size_average=True, val_range=None):
    if img2.dtype != img1.dtype:
        img2 = img2.to(dtype=img1.dtype)
    if img2.device != img1.device:
        img2 = img2.to(device=img1.device)

    if val_range is None:
        max_val = 255 if torch.max(img1) > 128 else 1
        min_val = -1 if torch.min(img1) < -0.5 else 0
        dynamic_range = max_val - min_val
    else:
        dynamic_range = val_range

    (_, channel, height, width) = img1.size()
    if window is None:
        real_size = min(window_size, height, width)
        window = create_window(real_size, channel=channel).to(img1.device).type(img1.dtype)
    else:
        window = window.to(img1.device).type(img1.dtype)

    mu1 = F.conv2d(img1, window, padding=0, groups=channel)
    mu2 = F.conv2d(img2, window, padding=0, groups=channel)
    mu1_sq = mu1.pow(2)
    mu2_sq = mu2.pow(2)
    mu1_mu2 = mu1 * mu2
    sigma1_sq = F.conv2d(img1 * img1, window, padding=0, groups=channel) - mu1_sq
    sigma2_sq = F.conv2d(img2 * img2, window, padding=0, groups=channel) - mu2_sq
    sigma12 = F.conv2d(img1 * img2, window, padding=0, groups=channel) - mu1_mu2
    c1 = (0.01 * dynamic_range) ** 2
    c2 = (0.03 * dynamic_range) ** 2
    ssim_map = ((2 * mu1_mu2 + c1) * (2.0 * sigma12 + c2)) / ((mu1_sq + mu2_sq + c1) * (sigma1_sq + sigma2_sq + c2))
    return ssim_map.mean() if size_average else ssim_map.mean(1).mean(1).mean(1)


def pack_prediction_for_forward(pred):
    times_probs = F.softmax(pred[:, 8:11], dim=1)
    return torch.cat([pred[:, :8], times_probs], dim=1)


def circular_deg_from_pair_np(sin_arr, cos_arr):
    ang = np.degrees(np.arctan2(sin_arr, cos_arr))
    return np.mod(ang, 360.0)


def decode_prediction_to_physical_np(pred_vec):
    pred_vec = np.asarray(pred_vec, dtype=np.float32)
    ia = np.clip(pred_vec[:, 0], 0.0, 1.0) * 90.0
    rad = np.clip(pred_vec[:, 1], 0.0, 1.0) * 200.0 + 50.0
    phi1 = circular_deg_from_pair_np(pred_vec[:, 2], pred_vec[:, 3])
    phi2 = circular_deg_from_pair_np(pred_vec[:, 4], pred_vec[:, 5])
    phi3 = circular_deg_from_pair_np(pred_vec[:, 6], pred_vec[:, 7])
    times_idx = np.argmax(pred_vec[:, 8:11], axis=1)
    times = times_idx.astype(np.float32) + 1.0
    phi2[times < 2] = 0.0
    phi3[times < 3] = 0.0
    raw = np.stack([ia, phi1, rad, times, phi2, phi3], axis=1).astype(np.float32)
    norm = np.stack([
        ia / 90.0,
        phi1 / 360.0,
        (rad - 50.0) / 200.0,
        times / 3.0,
        phi2 / 360.0,
        phi3 / 360.0,
    ], axis=1).astype(np.float32)
    return raw, norm


def compute_samplewise_image_metrics(target_imgs: torch.Tensor, pred_imgs: torch.Tensor):
    target_np = target_imgs.detach().cpu().numpy().astype(np.float32)
    pred_np = pred_imgs.detach().cpu().numpy().astype(np.float32)
    pcc_list = []
    ssim_list = []
    for i in range(target_np.shape[0]):
        tgt = target_np[i].reshape(-1)
        pred = pred_np[i].reshape(-1)
        if float(np.std(tgt)) > 0.0 and float(np.std(pred)) > 0.0:
            pcc = float(np.corrcoef(tgt, pred)[0, 1])
        else:
            pcc = 0.0
        tgt_t = torch.from_numpy(target_np[i:i + 1])
        pred_t = torch.from_numpy(pred_np[i:i + 1])
        ssim_val = float(ssim(pred_t, tgt_t, size_average=True).item())
        pcc_list.append(pcc)
        ssim_list.append(ssim_val)
    return np.asarray(pcc_list, dtype=np.float32), np.asarray(ssim_list, dtype=np.float32)


def bootstrap_confidence_interval(values, n_boot=1000, alpha=0.05, seed=42):
    arr = np.asarray(values, dtype=np.float64)
    arr = arr[np.isfinite(arr)]
    if arr.size == 0:
        return np.nan, np.nan
    rng = np.random.RandomState(seed)
    idx = rng.randint(0, arr.size, size=(n_boot, arr.size))
    means = arr[idx].mean(axis=1)
    lo = float(np.percentile(means, 100.0 * (alpha / 2.0)))
    hi = float(np.percentile(means, 100.0 * (1.0 - alpha / 2.0)))
    return lo, hi


def build_test_metrics_summary(target_imgs, pred_imgs, true_norm6_np, pred_norm6_np, seed):
    pcc_vals, ssim_vals = compute_samplewise_image_metrics(target_imgs, pred_imgs)
    summary = {
        "n_test": int(len(pcc_vals)),
        "pcc_mean": float(np.mean(pcc_vals)),
        "pcc_std": float(np.std(pcc_vals, ddof=1)) if len(pcc_vals) > 1 else 0.0,
        "ssim_mean": float(np.mean(ssim_vals)),
        "ssim_std": float(np.std(ssim_vals, ddof=1)) if len(ssim_vals) > 1 else 0.0,
        "param_r2_overall": float(r2_score(true_norm6_np, pred_norm6_np)),
    }
    summary["pcc_ci95_low"], summary["pcc_ci95_high"] = bootstrap_confidence_interval(pcc_vals, seed=seed)
    summary["ssim_ci95_low"], summary["ssim_ci95_high"] = bootstrap_confidence_interval(ssim_vals, seed=seed)
    for idx, name in enumerate(["ia", "phi1", "rad", "times", "phi2", "phi3"]):
        try:
            summary[f"r2_{name}"] = float(r2_score(true_norm6_np[:, idx], pred_norm6_np[:, idx]))
        except Exception:
            summary[f"r2_{name}"] = np.nan
    return summary, pcc_vals, ssim_vals


def run_test_in_batches(model_inv, model_fwd, test_dataset, device, batch_size, num_workers):
    loader = Data.DataLoader(
        dataset=test_dataset,
        batch_size=batch_size,
        num_workers=num_workers,
        pin_memory=torch.cuda.is_available(),
        persistent_workers=num_workers > 0,
        shuffle=False,
    )
    model_inv.eval()
    model_fwd.eval()
    pred_list, img_list, stru_list = [], [], []
    physical_norm6_list, physical_raw6_list = [], []

    with torch.no_grad():
        for xb, tb in loader:
            xb = xb.to(device, non_blocking=True)
            outb = model_inv(xb)
            strub = model_fwd(pack_prediction_for_forward(outb))
            pred_list.append(outb.cpu())
            img_list.append(xb.cpu())
            stru_list.append(strub.cpu())
            physical_norm6_list.append(tb["physical_norm6"].cpu())
            physical_raw6_list.append(tb["physical_raw6"].cpu())

    return {
        "imgs": torch.cat(img_list, dim=0),
        "preds": torch.cat(pred_list, dim=0),
        "strus": torch.cat(stru_list, dim=0),
        "physical_norm6": torch.cat(physical_norm6_list, dim=0),
        "physical_raw6": torch.cat(physical_raw6_list, dim=0),
    }


def export_outputs(output_dir: Path, result, metrics_summary, sample_pcc_vals, sample_ssim_vals):
    output_dir.mkdir(parents=True, exist_ok=True)
    true_img_dir = output_dir / "true"
    pred_img_dir = output_dir / "pred"
    result_dir = output_dir / "result"
    for d in [true_img_dir, pred_img_dir, result_dir]:
        d.mkdir(parents=True, exist_ok=True)

    to_pil = transforms.ToPILImage()
    save_n = min(100, len(result["imgs"]))
    for i in range(save_n):
        to_pil(result["imgs"][i]).save(str(true_img_dir / f"{i}.png"))
        to_pil(result["strus"][i]).save(str(pred_img_dir / f"{i}.png"))

    pred_vec_np = result["preds"].detach().numpy()
    true_norm6_np = result["physical_norm6"].detach().numpy()
    true_raw6_np = result["physical_raw6"].detach().numpy()
    pred_raw6_np, pred_norm6_np = decode_prediction_to_physical_np(pred_vec_np)

    pred_inc_angle = np.round(pred_raw6_np[:, 0])
    pred_azi_angle1 = np.round(pred_raw6_np[:, 1])
    pred_rad = np.round(pred_raw6_np[:, 2])
    pred_times = np.round(pred_raw6_np[:, 3])
    pred_azi_angle2 = np.round(pred_raw6_np[:, 4])
    pred_azi_angle3 = np.round(pred_raw6_np[:, 5])

    true_inc_angle = np.round(true_raw6_np[:, 0])
    true_azi_angle1 = np.round(true_raw6_np[:, 1])
    true_rad = np.round(true_raw6_np[:, 2])
    true_times = np.round(true_raw6_np[:, 3])
    true_azi_angle2 = np.round(true_raw6_np[:, 4])
    true_azi_angle3 = np.round(true_raw6_np[:, 5])

    np.savetxt(str(result_dir / f"pred_inc_angle-{MODEL_TAG}.txt"), pred_inc_angle, fmt="%d")
    np.savetxt(str(result_dir / f"true_inc_angle-{MODEL_TAG}.txt"), true_inc_angle, fmt="%d")
    np.savetxt(str(result_dir / f"pred_azi_angle1-{MODEL_TAG}.txt"), pred_azi_angle1, fmt="%d")
    np.savetxt(str(result_dir / f"true_azi_angle1-{MODEL_TAG}.txt"), true_azi_angle1, fmt="%d")
    np.savetxt(str(result_dir / f"pred_azi_angle2-{MODEL_TAG}.txt"), pred_azi_angle2, fmt="%d")
    np.savetxt(str(result_dir / f"true_azi_angle2-{MODEL_TAG}.txt"), true_azi_angle2, fmt="%d")
    np.savetxt(str(result_dir / f"pred_azi_angle3-{MODEL_TAG}.txt"), pred_azi_angle3, fmt="%d")
    np.savetxt(str(result_dir / f"true_azi_angle3-{MODEL_TAG}.txt"), true_azi_angle3, fmt="%d")
    np.savetxt(str(result_dir / f"pred_rad-{MODEL_TAG}.txt"), pred_rad, fmt="%d")
    np.savetxt(str(result_dir / f"true_rad-{MODEL_TAG}.txt"), true_rad, fmt="%d")
    np.savetxt(str(result_dir / f"pred_times-{MODEL_TAG}.txt"), pred_times, fmt="%d")
    np.savetxt(str(result_dir / f"true_times-{MODEL_TAG}.txt"), true_times, fmt="%d")
    np.savetxt(str(result_dir / f"test_pcc_samplewise-{MODEL_TAG}.txt"), sample_pcc_vals, fmt="%.8f")
    np.savetxt(str(result_dir / f"test_ssim_samplewise-{MODEL_TAG}.txt"), sample_ssim_vals, fmt="%.8f")

    pd.DataFrame([metrics_summary]).to_csv(result_dir / f"test_metrics_summary-{MODEL_TAG}.csv", index=False, encoding="utf-8-sig")
    with open(result_dir / f"test_metrics_summary-{MODEL_TAG}.json", "w", encoding="utf-8") as f:
        json.dump(metrics_summary, f, indent=2)

    m = min(100, len(pred_inc_angle))
    plt.figure(figsize=(12, 10))
    plt.suptitle("Test")
    plt.subplot(3, 2, 1)
    plt.plot(np.arange(1, m + 1), pred_inc_angle[:m], label="Inc Pred")
    plt.plot(np.arange(1, m + 1), true_inc_angle[:m], label="Inc True")
    plt.legend()
    plt.subplot(3, 2, 2)
    plt.plot(np.arange(1, m + 1), pred_azi_angle1[:m], label="Azi1 Pred")
    plt.plot(np.arange(1, m + 1), true_azi_angle1[:m], label="Azi1 True")
    plt.legend()
    plt.subplot(3, 2, 3)
    plt.plot(np.arange(1, m + 1), pred_rad[:m], label="Rad Pred")
    plt.plot(np.arange(1, m + 1), true_rad[:m], label="Rad True")
    plt.legend()
    plt.subplot(3, 2, 4)
    plt.plot(np.arange(1, m + 1), pred_azi_angle2[:m], label="Azi2 Pred")
    plt.plot(np.arange(1, m + 1), true_azi_angle2[:m], label="Azi2 True")
    plt.legend()
    plt.subplot(3, 2, 5)
    plt.plot(np.arange(1, m + 1), pred_azi_angle3[:m], label="Azi3 Pred")
    plt.plot(np.arange(1, m + 1), true_azi_angle3[:m], label="Azi3 True")
    plt.legend()
    plt.subplot(3, 2, 6)
    plt.plot(np.arange(1, m + 1), pred_times[:m], label="Times Pred")
    plt.plot(np.arange(1, m + 1), true_times[:m], label="Times True")
    plt.legend()
    plt.tight_layout()
    plt.savefig(str(result_dir / f"prediction_curves-{MODEL_TAG}.png"), dpi=300)
    plt.close()

    if openpyxl is not None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "sheet1"
        headers = (
            "True-inc", "Pred-inc", "True-azi1", "Pred-azi1", "True-rad", "Pred-rad",
            "True-azi2", "Pred-azi2", "True-azi3", "Pred-azi3", "True-times", "Pred-times",
        )
        for idx, item in enumerate(headers):
            ws.cell(row=1, column=idx + 1, value=item)
        for i in range(m):
            ws.cell(row=i + 2, column=1, value=float(true_inc_angle[i]))
            ws.cell(row=i + 2, column=2, value=float(pred_inc_angle[i]))
            ws.cell(row=i + 2, column=3, value=float(true_azi_angle1[i]))
            ws.cell(row=i + 2, column=4, value=float(pred_azi_angle1[i]))
            ws.cell(row=i + 2, column=5, value=float(true_rad[i]))
            ws.cell(row=i + 2, column=6, value=float(pred_rad[i]))
            ws.cell(row=i + 2, column=7, value=float(true_azi_angle2[i]))
            ws.cell(row=i + 2, column=8, value=float(pred_azi_angle2[i]))
            ws.cell(row=i + 2, column=9, value=float(true_azi_angle3[i]))
            ws.cell(row=i + 2, column=10, value=float(pred_azi_angle3[i]))
            ws.cell(row=i + 2, column=11, value=float(true_times[i]))
            ws.cell(row=i + 2, column=12, value=float(pred_times[i]))
        wb.save(str(result_dir / f"test-pred-paramas-{MODEL_TAG}.xlsx"))


def main():
    args = parse_args()
    script_dir = Path(__file__).resolve().parent
    model_dir = (script_dir / args.model_dir).resolve() if not args.model_dir.is_absolute() else args.model_dir
    output_dir = (script_dir / args.output_dir).resolve() if not args.output_dir.is_absolute() else args.output_dir

    set_seed(args.seed)
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    transform = transforms.Compose([
        transforms.Resize(SOURCE_IMAGE_SIZE),
        transforms.Grayscale(num_output_channels=IMAGE_CHANNELS),
        transforms.ToTensor(),
    ])

    samples = build_samples(args.data_root, args.excel_path)
    rng = np.random.RandomState(args.seed)
    index = np.arange(len(samples))
    rng.shuffle(index)

    n = len(index)
    n_train = int(n * 0.7)
    n_val = int(n * 0.2)
    test_idx = index[n_train + n_val:].tolist()
    test_samples = [samples[int(i)] for i in test_idx]
    selected_test_samples = stratified_select_test_subset(test_samples, args.test_sample_size, transform, seed=args.seed)
    test_set = LazyImageLabelDataset(selected_test_samples, transform=transform)

    model_inverse = InverseBiCBAMNetV2(in_channels=IMAGE_CHANNELS, ratio=8).to(device)
    model_forward = ForwardBiCBAMNetV2(out_channels=IMAGE_CHANNELS, img_size=SOURCE_IMAGE_SIZE, ratio=8).to(device)

    ckpt_inverse = model_dir / f"model_inverse.pth"
    ckpt_forward = model_dir / f"model_forward.pth"
    if (not ckpt_inverse.exists()) or (not ckpt_forward.exists()):
        raise FileNotFoundError(f"Missing checkpoints:\n{ckpt_inverse}\n{ckpt_forward}")
    model_inverse.load_state_dict(torch.load(str(ckpt_inverse), map_location=device))
    model_forward.load_state_dict(torch.load(str(ckpt_forward), map_location=device))

    result = run_test_in_batches(model_inverse, model_forward, test_set, device, args.batch_size, args.num_workers)
    pred_vec_np = result["preds"].detach().numpy()
    true_norm6_np = result["physical_norm6"].detach().numpy()
    _, pred_norm6_np = decode_prediction_to_physical_np(pred_vec_np)
    metrics_summary, sample_pcc_vals, sample_ssim_vals = build_test_metrics_summary(
        result["imgs"], result["strus"], true_norm6_np, pred_norm6_np, args.seed
    )
    export_outputs(output_dir, result, metrics_summary, sample_pcc_vals, sample_ssim_vals)
    print(json.dumps(metrics_summary, indent=2))


if __name__ == "__main__":
    main()
